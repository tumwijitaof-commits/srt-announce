from flask import (Flask, request, render_template_string, send_from_directory, jsonify,
                   session, redirect, url_for, flash, send_file, abort, g)
from pathlib import Path
import os
import subprocess
import json
import time
import uuid
import re
import hashlib
import threading
import atexit
import sqlite3
try:
    import psycopg
except ImportError:
    psycopg = None
try:
    from psycopg_pool import ConnectionPool
except ImportError:
    ConnectionPool = None
import csv
import io
import shutil
from datetime import datetime, date, time as dt_time
from functools import wraps
from zoneinfo import ZoneInfo
from werkzeug.security import generate_password_hash, check_password_hash
from concurrent.futures import ThreadPoolExecutor

app = Flask(__name__)
BASE_DIR = Path(__file__).resolve().parent

# รหัสลับของ session เก็บถาวรข้างไฟล์โปรแกรม เพื่อให้ login ไม่หลุดเมื่อรีสตาร์ต
_SECRET_FILE = BASE_DIR / ".station_session_secret"
if os.environ.get("SECRET_KEY"):
    app.secret_key = os.environ["SECRET_KEY"]
elif _SECRET_FILE.exists():
    app.secret_key = _SECRET_FILE.read_text(encoding="utf-8").strip()
else:
    app.secret_key = uuid.uuid4().hex + uuid.uuid4().hex
    try:
        _SECRET_FILE.write_text(app.secret_key, encoding="utf-8")
    except OSError:
        pass
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE="Lax")

BANGKOK_TZ = ZoneInfo("Asia/Bangkok")
DB_PATH = BASE_DIR / "station_system.db"
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = "postgresql://" + DATABASE_URL[len("postgres://"):]
# Supabase requires encrypted PostgreSQL connections. Add sslmode only when absent.
if USE_POSTGRES and "supabase" in DATABASE_URL and "sslmode=" not in DATABASE_URL:
    DATABASE_URL += ("&" if "?" in DATABASE_URL else "?") + "sslmode=require"

# ค่าประสิทธิภาพฐานข้อมูล: ใช้ connection เดิมซ้ำ ลดเวลาจับมือกับ Supabase ทุกครั้งที่กดเมนู
POSTGRES_POOL_MIN_SIZE = max(1, int(os.environ.get("POSTGRES_POOL_MIN_SIZE", "1")))
POSTGRES_POOL_MAX_SIZE = max(POSTGRES_POOL_MIN_SIZE, int(os.environ.get("POSTGRES_POOL_MAX_SIZE", "3")))
POSTGRES_POOL_TIMEOUT = max(3, int(os.environ.get("POSTGRES_POOL_TIMEOUT", "10")))
USER_SESSION_CACHE_SECONDS = max(15, int(os.environ.get("USER_SESSION_CACHE_SECONDS", "90")))
TRAIN_CACHE_SECONDS = max(0, int(os.environ.get("TRAIN_CACHE_SECONDS", "30")))

AUDIO_DIR = BASE_DIR / "audio_generated"
AUDIO_DIR.mkdir(exist_ok=True)

THAI_VOICE_OPTIONS = {
    "th-TH-PremwadeeNeural": "เสียงหญิง — ชัดเจน เป็นธรรมชาติ",
    "th-TH-NiwatNeural": "เสียงชาย — สุภาพ เป็นทางการ",
}
VOICE_NAME = os.environ.get("TTS_VOICE", "th-TH-PremwadeeNeural")
if VOICE_NAME not in THAI_VOICE_OPTIONS:
    VOICE_NAME = "th-TH-PremwadeeNeural"

# ไม่เร่งความดังหรือกดระดับเสียงมากเกินไป เพราะจะทำให้เสียงแตกและคำเพี้ยน
TTS_RATE = os.environ.get("TTS_RATE", "-10%")
TTS_VOLUME = os.environ.get("TTS_VOLUME", "+0%")
TTS_PITCH = os.environ.get("TTS_PITCH", "+0Hz")
ENGLISH_VOICE_OPTIONS = {
    "female": os.environ.get("TTS_EN_FEMALE_VOICE", "en-US-JennyNeural"),
    "male": os.environ.get("TTS_EN_MALE_VOICE", "en-US-GuyNeural"),
}
TTS_EN_RATE = os.environ.get("TTS_EN_RATE", "-8%")
TTS_EN_VOLUME = os.environ.get("TTS_EN_VOLUME", "+0%")
TTS_EN_PITCH = os.environ.get("TTS_EN_PITCH", "+0Hz")
STATION_NAME = "คลองบางพระ"
CHIME_FILENAME = "chime.mp3"

# แคชไฟล์เสียงและล็อกสำหรับป้องกันการสร้างเสียงซ้ำพร้อมกัน
_AUDIO_CACHE_LOCKS = {}
_AUDIO_CACHE_LOCKS_GUARD = threading.Lock()
_AUDIO_CLEANUP_LOCK = threading.Lock()
_LAST_AUDIO_CLEANUP = 0.0
AUDIO_CACHE_MAX_AGE = int(os.environ.get("AUDIO_CACHE_MAX_AGE", 7 * 24 * 60 * 60))


# ------------------------------------------------------------
# ฐานข้อมูลตารางเดินรถ สถานีคลองบางพระ
# ------------------------------------------------------------
INBOUND_TRAINS = [
    {"label": "384 (05:30) ฉะเชิงเทรา - กรุงเทพ", "num": "384", "origin": "ชุมทางฉะเชิงเทรา", "dest": "กรุงเทพ", "time": "5 นาฬิกา 30 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "380 (05:55) ฉะเชิงเทรา - กรุงเทพ", "num": "380", "origin": "ชุมทางฉะเชิงเทรา", "dest": "กรุงเทพ", "time": "5 นาฬิกา 55 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "372 (06:30) ปราจีนบุรี - กรุงเทพ", "num": "372", "origin": "ปราจีนบุรี", "dest": "กรุงเทพ", "time": "6 นาฬิกา 30 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "388 (07:12) ฉะเชิงเทรา - กรุงเทพ", "num": "388", "origin": "ชุมทางฉะเชิงเทรา", "dest": "กรุงเทพ", "time": "7 นาฬิกา 12 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "278 (08:41) กบินทร์บุรี - กรุงเทพ", "num": "278", "origin": "กบินทร์บุรี", "dest": "กรุงเทพ", "time": "8 นาฬิกา 41 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "280 (10:33) คลองลึก - กรุงเทพ", "num": "280", "origin": "ด่านพรมแดนบ้านคลองลึก", "dest": "กรุงเทพ", "time": "10 นาฬิกา 33 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "368 (12:44) ฉะเชิงเทรา - กรุงเทพ", "num": "368", "origin": "ชุมทางฉะเชิงเทรา", "dest": "กรุงเทพ", "time": "12 นาฬิกา 44 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "390 (14:12) ฉะเชิงเทรา - กรุงเทพ", "num": "390", "origin": "ชุมทางฉะเชิงเทรา", "dest": "กรุงเทพ", "time": "14 นาฬิกา 12 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "282 (15:42) กบินทร์บุรี - กรุงเทพ", "num": "282", "origin": "กบินทร์บุรี", "dest": "กรุงเทพ", "time": "15 นาฬิกา 42 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "284 (16:33) จุกเสม็ด - กรุงเทพ", "num": "284", "origin": "จุกเสม็ด", "dest": "กรุงเทพ", "time": "16 นาฬิกา 33 นาที", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
    {"label": "276 (19:00) คลองลึก - กรุงเทพ", "num": "276", "origin": "ด่านพรมแดนบ้านคลองลึก", "dest": "กรุงเทพ", "time": "19 นาฬิกา", "next": "ป้ายหยุดรถคลองแขวงกลั่น และ สถานีคลองเปรง"},
]

OUTBOUND_TRAINS = [
    {"label": "275 (07:28) กรุงเทพ - คลองลึก", "num": "275", "origin": "กรุงเทพ", "dest": "ด่านพรมแดนบ้านคลองลึก", "time": "7 นาฬิกา 28 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "283 (08:46) กรุงเทพ - จุกเสม็ด", "num": "283", "origin": "กรุงเทพ", "dest": "จุกเสม็ด", "time": "8 นาฬิกา 46 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "281 (09:23) กรุงเทพ - กบินทร์บุรี", "num": "281", "origin": "กรุงเทพ", "dest": "กบินทร์บุรี", "time": "9 นาฬิกา 23 นาที", "next": "สถานีชุมทางฉะเชิงเทรา"},
    {"label": "367 (11:35) กรุงเทพ - ฉะเชิงเทรา", "num": "367", "origin": "กรุงเทพ", "dest": "ชุมทางฉะเชิงเทรา", "time": "11 นาฬิกา 35 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "389 (13:23) กรุงเทพ - ฉะเชิงเทรา", "num": "389", "origin": "กรุงเทพ", "dest": "ชุมทางฉะเชิงเทรา", "time": "13 นาฬิกา 23 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "279 (14:08) กรุงเทพ - คลองลึก", "num": "279", "origin": "กรุงเทพ", "dest": "ด่านพรมแดนบ้านคลองลึก", "time": "14 นาฬิกา 8 นาที", "next": "สถานีชุมทางฉะเชิงเทรา"},
    {"label": "277 (16:37) กรุงเทพ - กบินทร์บุรี", "num": "277", "origin": "กรุงเทพ", "dest": "กบินทร์บุรี", "time": "16 นาฬิกา 37 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "379 (17:57) กรุงเทพ - ฉะเชิงเทรา", "num": "379", "origin": "กรุงเทพ", "dest": "ชุมทางฉะเชิงเทรา", "time": "17 นาฬิกา 57 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "391 (18:17) กรุงเทพ - ฉะเชิงเทรา", "num": "391", "origin": "กรุงเทพ", "dest": "ชุมทางฉะเชิงเทรา", "time": "18 นาฬิกา 17 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "371 (19:11) กรุงเทพ - ปราจีนบุรี", "num": "371", "origin": "กรุงเทพ", "dest": "ปราจีนบุรี", "time": "19 นาฬิกา 11 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
    {"label": "383 (20:25) กรุงเทพ - ฉะเชิงเทรา", "num": "383", "origin": "กรุงเทพ", "dest": "ชุมทางฉะเชิงเทรา", "time": "20 นาฬิกา 25 นาที", "next": "ป้ายหยุดรถบางเตย และ สถานีชุมทางฉะเชิงเทรา"},
]

TRAIN_DATA = {train["label"]: train for train in INBOUND_TRAINS + OUTBOUND_TRAINS}

ANNOUNCEMENT_BUTTONS = [
    {"idx": 0, "title": "ขอทาง / ขายตั๋ว", "hint": "แจ้งผู้โดยสารให้ซื้อตั๋วก่อนเดินทาง", "group": "ก่อนรถเข้า"},
    {"idx": 1, "title": "รอรับโดยสาร", "hint": "ให้ผู้โดยสารรอที่ชานชาลา", "group": "ก่อนรถเข้า"},
    {"idx": 2, "title": "รถกำลังเข้าเทียบ", "hint": "เตือนยืนหลังเส้นสีเหลือง", "group": "รถเข้า-ออก"},
    {"idx": 3, "title": "รถผ่านสถานี", "hint": "ประกาศรถผ่านขบวนปกติ", "group": "รถเข้า-ออก"},
    {"idx": 4, "title": "รถจอดรับส่ง / ออก", "hint": "ประกาศสถานีถัดไปและให้ขึ้นรถ", "group": "รถเข้า-ออก"},
    {"idx": 5, "title": "รถล่าช้า", "hint": "แจ้งเวลาคาดว่าจะถึง", "group": "เหตุการณ์พิเศษ"},
    {"idx": 6, "title": "ระวังคนลงรถ", "hint": "เตือนผู้โดยสารขณะรถเข้า", "group": "ความปลอดภัย"},
    {"idx": 7, "title": "ห้ามสูบบุหรี่", "hint": "ประกาศขอความร่วมมือ", "group": "ความปลอดภัย"},
    {"idx": 8, "title": "ประกาศเอง", "hint": "อ่านข้อความที่พิมพ์เอง", "group": "ประกาศทั่วไป"},
    {"idx": 9, "title": "สินค้า / พิเศษ ผ่าน", "hint": "เลือกประเภทรถวิ่งผ่าน", "group": "เหตุการณ์พิเศษ"},
    {"idx": 10, "title": "รถเข้าพร้อมกัน 2–3 ขบวน", "hint": "ใช้ข้อมูลขบวนที่ 1, 2 และขบวนที่ 3 ถ้ามี", "group": "เหตุการณ์พิเศษ"},
]

# ------------------------------------------------------------
# ระบบฐานข้อมูล / ผู้ใช้ / ตารางรถ / ประวัติ / ตรวจสุขภาพ
# เพิ่มแยกจากระบบ TTS เดิม เพื่อไม่เปลี่ยนจังหวะ เสียง หรือความเร็ว
# ------------------------------------------------------------
ROLE_LABELS = {
    "announcer": "เจ้าหน้าที่ประกาศ",
    "admin": "ผู้ดูแลระบบ",
    "auditor": "ผู้ตรวจสอบประวัติ",
}
SERVICE_LABELS = {
    "daily": "ทุกวัน",
    "weekday": "วันธรรมดา (จันทร์–ศุกร์)",
    "weekend": "วันเสาร์–อาทิตย์",
    "holiday": "วันหยุดตามวันที่ระบุ",
    "custom": "เฉพาะวันที่ระบุ",
}
DIRECTION_LABELS = {"inbound": "ขาเข้า กรุงเทพ", "outbound": "ขาออก ไปทางตะวันออก"}


def now_bangkok():
    return datetime.now(BANGKOK_TZ)


def now_iso():
    return now_bangkok().isoformat(timespec="seconds")


class CompatRow(dict):
    """แถวข้อมูลที่อ่านได้ทั้ง row["column"] และ row[0] เหมือน sqlite3.Row"""

    def __getitem__(self, key):
        if isinstance(key, int):
            return tuple(self.values())[key]
        return super().__getitem__(key)


class PostgresCursor:
    def __init__(self, cursor, lastrowid=None):
        self._cursor = cursor
        self.lastrowid = lastrowid

    def _wrap(self, row):
        if row is None:
            return None
        columns = [item.name if hasattr(item, "name") else item[0] for item in self._cursor.description]
        return CompatRow(zip(columns, row))

    def fetchone(self):
        return self._wrap(self._cursor.fetchone())

    def fetchall(self):
        rows = self._cursor.fetchall()
        if not rows:
            return []
        columns = [item.name if hasattr(item, "name") else item[0] for item in self._cursor.description]
        return [CompatRow(zip(columns, row)) for row in rows]


_POSTGRES_POOL = None
_POSTGRES_POOL_LOCK = threading.Lock()


def get_postgres_pool():
    """สร้าง pool เพียงครั้งเดียวต่อ Gunicorn worker แล้วนำ connection กลับมาใช้ซ้ำ"""
    global _POSTGRES_POOL
    if _POSTGRES_POOL is not None:
        return _POSTGRES_POOL
    if psycopg is None or ConnectionPool is None:
        raise RuntimeError("ยังไม่มี psycopg pool กรุณาติดตั้ง psycopg[binary] และ psycopg_pool")
    with _POSTGRES_POOL_LOCK:
        if _POSTGRES_POOL is None:
            pool = ConnectionPool(
                conninfo=DATABASE_URL,
                min_size=POSTGRES_POOL_MIN_SIZE,
                max_size=POSTGRES_POOL_MAX_SIZE,
                timeout=POSTGRES_POOL_TIMEOUT,
                max_idle=300,
                max_lifetime=1800,
                reconnect_timeout=30,
                open=False,
                name="bangphra-db-pool",
            )
            # เปิด pool และเตรียม connection แรกตั้งแต่ตอนเริ่มระบบ แทนที่จะรอเมื่อผู้ใช้กดเมนูครั้งแรก
            pool.open()
            pool.wait(timeout=POSTGRES_POOL_TIMEOUT + 5)
            _POSTGRES_POOL = pool
    return _POSTGRES_POOL


def _close_postgres_pool():
    global _POSTGRES_POOL
    pool = _POSTGRES_POOL
    _POSTGRES_POOL = None
    if pool is not None:
        try:
            pool.close()
        except Exception:
            pass


atexit.register(_close_postgres_pool)


class PostgresConnection:
    """ตัวแปลงคำสั่ง SQLite เดิมให้ทำงานกับ PostgreSQL ผ่าน connection pool"""

    def __init__(self, database_url):
        self._database_url = database_url
        self._pool_context = None
        self._conn = None

    def __enter__(self):
        # pool.connection() จะ commit/rollback และคืน connection ให้ pool ให้อัตโนมัติ
        self._pool_context = get_postgres_pool().connection(timeout=POSTGRES_POOL_TIMEOUT)
        self._conn = self._pool_context.__enter__()
        return self

    def __exit__(self, exc_type, exc, traceback):
        if self._pool_context is None:
            return False
        try:
            return self._pool_context.__exit__(exc_type, exc, traceback)
        finally:
            self._conn = None
            self._pool_context = None

    @staticmethod
    def _convert_sql(sql):
        # SQLite ใช้ ? แต่ psycopg ใช้ %s
        sql = re.sub(
            r"username\s*=\s*\?\s+COLLATE\s+NOCASE",
            "LOWER(username)=LOWER(%s)",
            sql,
            flags=re.IGNORECASE,
        )
        sql = re.sub(r"\s+COLLATE\s+NOCASE", "", sql, flags=re.IGNORECASE)
        return sql.replace("?", "%s")

    def execute(self, sql, params=()):
        if self._conn is None:
            raise RuntimeError("ยังไม่ได้เปิดการเชื่อมต่อฐานข้อมูล")
        converted = self._convert_sql(sql)
        cursor = self._conn.cursor()
        # เฉพาะ INSERT ที่โค้ดเดิมอ่าน lastrowid
        insert_match = re.match(r"\s*INSERT\s+INTO\s+(schedule_versions|announcement_history)\b", converted, re.I)
        needs_id = bool(insert_match and " returning " not in converted.lower() and " select " not in converted.lower())
        if needs_id:
            converted = converted.rstrip().rstrip(";") + " RETURNING id"
        cursor.execute(converted, tuple(params or ()))
        lastrowid = None
        if needs_id:
            returned = cursor.fetchone()
            lastrowid = returned[0] if returned else None
        return PostgresCursor(cursor, lastrowid=lastrowid)

    def executemany(self, sql, rows):
        if self._conn is None:
            raise RuntimeError("ยังไม่ได้เปิดการเชื่อมต่อฐานข้อมูล")
        cursor = self._conn.cursor()
        cursor.executemany(self._convert_sql(sql), rows)
        return PostgresCursor(cursor)

    def executescript(self, script):
        if self._conn is None:
            raise RuntimeError("ยังไม่ได้เปิดการเชื่อมต่อฐานข้อมูล")
        cursor = self._conn.cursor()
        for statement in script.split(";"):
            statement = statement.strip()
            if statement:
                cursor.execute(statement)
        return PostgresCursor(cursor)


def get_db():
    if USE_POSTGRES:
        return PostgresConnection(DATABASE_URL)
    conn = sqlite3.connect(DB_PATH, timeout=20)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn


SQLITE_SCHEMA = r"""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL UNIQUE COLLATE NOCASE,
    password_hash TEXT NOT NULL,
    display_name TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN ('announcer','admin','auditor')),
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    last_login TEXT
);
CREATE TABLE IF NOT EXISTS schedule_versions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    effective_date TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'published' CHECK(status IN ('draft','published')),
    created_at TEXT NOT NULL,
    created_by INTEGER,
    FOREIGN KEY(created_by) REFERENCES users(id)
);
CREATE TABLE IF NOT EXISTS trains (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    version_id INTEGER NOT NULL,
    direction TEXT NOT NULL CHECK(direction IN ('inbound','outbound')),
    num TEXT NOT NULL,
    label TEXT NOT NULL,
    origin TEXT NOT NULL,
    dest TEXT NOT NULL,
    time_hhmm TEXT NOT NULL,
    time_spoken TEXT NOT NULL,
    next_station TEXT NOT NULL DEFAULT '',
    service_pattern TEXT NOT NULL DEFAULT 'daily' CHECK(service_pattern IN ('daily','weekday','weekend','holiday','custom')),
    service_dates TEXT NOT NULL DEFAULT '',
    enabled INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(version_id) REFERENCES schedule_versions(id) ON DELETE CASCADE,
    UNIQUE(version_id, direction, num, time_hhmm)
);
CREATE TABLE IF NOT EXISTS announcement_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    started_at TEXT NOT NULL,
    user_id INTEGER,
    username TEXT NOT NULL,
    train_num TEXT,
    announcement_type TEXT NOT NULL,
    announce_mode TEXT NOT NULL,
    voice TEXT NOT NULL,
    platform TEXT,
    message TEXT,
    generation_ms INTEGER,
    playback_success INTEGER,
    pause_times TEXT NOT NULL DEFAULT '[]',
    stop_time TEXT,
    completed_at TEXT,
    failure_reason TEXT,
    FOREIGN KEY(user_id) REFERENCES users(id)
);
CREATE TABLE IF NOT EXISTS announcement_events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    history_id INTEGER NOT NULL,
    event_type TEXT NOT NULL,
    event_at TEXT NOT NULL,
    details TEXT,
    FOREIGN KEY(history_id) REFERENCES announcement_history(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_trains_version ON trains(version_id, direction, time_hhmm);
CREATE INDEX IF NOT EXISTS idx_history_started ON announcement_history(started_at DESC);
CREATE INDEX IF NOT EXISTS idx_events_history ON announcement_events(history_id, event_at);
"""

POSTGRES_SCHEMA = r"""
CREATE TABLE IF NOT EXISTS users (
    id BIGSERIAL PRIMARY KEY,
    username TEXT NOT NULL,
    password_hash TEXT NOT NULL,
    display_name TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN ('announcer','admin','auditor')),
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    last_login TEXT
);
CREATE UNIQUE INDEX IF NOT EXISTS uq_users_username_lower ON users (LOWER(username));
CREATE TABLE IF NOT EXISTS schedule_versions (
    id BIGSERIAL PRIMARY KEY,
    name TEXT NOT NULL,
    effective_date TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'published' CHECK(status IN ('draft','published')),
    created_at TEXT NOT NULL,
    created_by BIGINT REFERENCES users(id)
);
CREATE TABLE IF NOT EXISTS trains (
    id BIGSERIAL PRIMARY KEY,
    version_id BIGINT NOT NULL REFERENCES schedule_versions(id) ON DELETE CASCADE,
    direction TEXT NOT NULL CHECK(direction IN ('inbound','outbound')),
    num TEXT NOT NULL,
    label TEXT NOT NULL,
    origin TEXT NOT NULL,
    dest TEXT NOT NULL,
    time_hhmm TEXT NOT NULL,
    time_spoken TEXT NOT NULL,
    next_station TEXT NOT NULL DEFAULT '',
    service_pattern TEXT NOT NULL DEFAULT 'daily' CHECK(service_pattern IN ('daily','weekday','weekend','holiday','custom')),
    service_dates TEXT NOT NULL DEFAULT '',
    enabled INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE(version_id, direction, num, time_hhmm)
);
CREATE TABLE IF NOT EXISTS announcement_history (
    id BIGSERIAL PRIMARY KEY,
    started_at TEXT NOT NULL,
    user_id BIGINT REFERENCES users(id),
    username TEXT NOT NULL,
    train_num TEXT,
    announcement_type TEXT NOT NULL,
    announce_mode TEXT NOT NULL,
    voice TEXT NOT NULL,
    platform TEXT,
    message TEXT,
    generation_ms INTEGER,
    playback_success INTEGER,
    pause_times TEXT NOT NULL DEFAULT '[]',
    stop_time TEXT,
    completed_at TEXT,
    failure_reason TEXT
);
CREATE TABLE IF NOT EXISTS announcement_events (
    id BIGSERIAL PRIMARY KEY,
    history_id BIGINT NOT NULL REFERENCES announcement_history(id) ON DELETE CASCADE,
    event_type TEXT NOT NULL,
    event_at TEXT NOT NULL,
    details TEXT
);
CREATE INDEX IF NOT EXISTS idx_trains_version ON trains(version_id, direction, time_hhmm);
CREATE INDEX IF NOT EXISTS idx_history_started ON announcement_history(started_at DESC);
CREATE INDEX IF NOT EXISTS idx_events_history ON announcement_events(history_id, event_at);
"""

DB_INTEGRITY_ERRORS = (sqlite3.IntegrityError,)
if psycopg is not None:
    DB_INTEGRITY_ERRORS = DB_INTEGRITY_ERRORS + (psycopg.IntegrityError,)

def _label_time(label):
    match = re.search(r"\((\d{1,2}:\d{2})\)", label or "")
    return match.group(1) if match else "00:00"


def spoken_time_from_hhmm(value):
    value = (value or "").strip()
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", value)
    if not match:
        return value
    hour = int(match.group(1))
    minute = int(match.group(2))
    if minute == 0:
        return f"{hour} นาฬิกา"
    return f"{hour} นาฬิกา {minute} นาที"


def normalize_date_list(value):
    raw = value or ""
    parts = re.split(r"[,;\n\s]+", raw.strip()) if raw.strip() else []
    normalized = []
    for item in parts:
        if not item:
            continue
        try:
            normalized.append(date.fromisoformat(item).isoformat())
        except ValueError:
            raise ValueError(f"วันที่ {item} ไม่ถูกต้อง ต้องเป็นรูปแบบ YYYY-MM-DD")
    return ",".join(sorted(set(normalized)))


def init_database():
    with get_db() as conn:
        conn.executescript(POSTGRES_SCHEMA if USE_POSTGRES else SQLITE_SCHEMA)

        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            initial_password = os.environ.get("INITIAL_ADMIN_PASSWORD", "admin1234")
            conn.execute(
                "INSERT INTO users(username,password_hash,display_name,role,active,created_at) VALUES(?,?,?,?,1,?)",
                ("admin", generate_password_hash(initial_password), "ผู้ดูแลระบบ", "admin", now_iso()),
            )

        if conn.execute("SELECT COUNT(*) FROM schedule_versions").fetchone()[0] == 0:
            admin_id = conn.execute("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1").fetchone()[0]
            cursor = conn.execute(
                "INSERT INTO schedule_versions(name,effective_date,status,created_at,created_by) VALUES(?,?,?,?,?)",
                ("ตารางเริ่มต้นจากโค้ดเดิม", "2026-01-01", "published", now_iso(), admin_id),
            )
            version_id = cursor.lastrowid
            rows = []
            for direction, trains in (("inbound", INBOUND_TRAINS), ("outbound", OUTBOUND_TRAINS)):
                for train in trains:
                    hhmm = _label_time(train["label"])
                    rows.append((
                        version_id, direction, train["num"], train["label"], train["origin"], train["dest"],
                        hhmm, train["time"], train.get("next", ""), "daily", "", 1, now_iso(), now_iso(),
                    ))
            conn.executemany(
                """INSERT INTO trains(version_id,direction,num,label,origin,dest,time_hhmm,time_spoken,next_station,
                   service_pattern,service_dates,enabled,created_at,updated_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                rows,
            )


def _store_session_user(user):
    profile = {
        "id": int(user["id"]),
        "username": user["username"],
        "display_name": user["display_name"],
        "role": user["role"],
        "active": int(user["active"]),
    }
    session["user_id"] = profile["id"]
    session["user_profile"] = profile
    session["user_checked_at"] = time.time()
    g._current_user = profile
    return profile


def _invalidate_session_user_cache():
    session.pop("user_profile", None)
    session.pop("user_checked_at", None)
    if hasattr(g, "_current_user"):
        delattr(g, "_current_user")


def get_current_user():
    # ภายใน request เดียวกัน decorator และหน้าเว็บใช้ข้อมูลเดียวกัน ไม่ยิง query ซ้ำ
    if hasattr(g, "_current_user"):
        return g._current_user

    user_id = session.get("user_id")
    if not user_id:
        g._current_user = None
        return None

    cached = session.get("user_profile")
    checked_at = float(session.get("user_checked_at") or 0)
    if (
        isinstance(cached, dict)
        and cached.get("id") == user_id
        and cached.get("active")
        and time.time() - checked_at < USER_SESSION_CACHE_SECONDS
    ):
        g._current_user = dict(cached)
        return g._current_user

    with get_db() as conn:
        row = conn.execute(
            "SELECT id,username,display_name,role,active FROM users WHERE id=?", (user_id,)
        ).fetchone()
    if not row or not row["active"]:
        session.clear()
        g._current_user = None
        return None
    return _store_session_user(row)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not get_current_user():
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def roles_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect(url_for("login", next=request.path))
            if user["role"] not in roles:
                abort(403)
            return view(*args, **kwargs)
        return wrapped
    return decorator


_TRAIN_LIST_CACHE = {}
_TRAIN_LIST_CACHE_LOCK = threading.Lock()


def invalidate_train_cache():
    with _TRAIN_LIST_CACHE_LOCK:
        _TRAIN_LIST_CACHE.clear()


def _clone_train_result(result):
    inbound, outbound, data, version = result
    return (
        [dict(item) for item in inbound],
        [dict(item) for item in outbound],
        {key: dict(item) for key, item in data.items()},
        dict(version) if version else None,
    )


def version_for_date(target_date=None, conn=None):
    target_date = target_date or now_bangkok().date()

    def _read(active_conn):
        row = active_conn.execute(
            """SELECT * FROM schedule_versions
               WHERE status='published' AND effective_date<=?
               ORDER BY effective_date DESC, id DESC LIMIT 1""",
            (target_date.isoformat(),),
        ).fetchone()
        if not row:
            row = active_conn.execute(
                "SELECT * FROM schedule_versions WHERE status='published' ORDER BY effective_date ASC,id ASC LIMIT 1"
            ).fetchone()
        return dict(row) if row else None

    if conn is not None:
        return _read(conn)
    with get_db() as db_conn:
        return _read(db_conn)


def train_operates_on(row, target_date):
    pattern = row["service_pattern"]
    weekday = target_date.weekday()
    if pattern == "daily":
        return True
    if pattern == "weekday":
        return weekday < 5
    if pattern == "weekend":
        return weekday >= 5
    dates = {item for item in (row["service_dates"] or "").split(",") if item}
    return target_date.isoformat() in dates


def get_active_train_lists(target_date=None):
    target_date = target_date or now_bangkok().date()
    cache_key = target_date.isoformat()
    if TRAIN_CACHE_SECONDS:
        with _TRAIN_LIST_CACHE_LOCK:
            cached = _TRAIN_LIST_CACHE.get(cache_key)
            if cached and time.monotonic() - cached[0] < TRAIN_CACHE_SECONDS:
                return _clone_train_result(cached[1])

    # อ่าน version และขบวนด้วย connection เดียว ลดรอบการเดินทาง Virginia ↔ Supabase
    with get_db() as conn:
        version = version_for_date(target_date, conn=conn)
        if not version:
            result = ([], [], {}, None)
        else:
            rows = conn.execute(
                "SELECT * FROM trains WHERE version_id=? AND enabled=1 ORDER BY direction,time_hhmm,num",
                (version["id"],),
            ).fetchall()
            inbound, outbound = [], []
            for row in rows:
                if not train_operates_on(row, target_date):
                    continue
                item = {
                    "label": row["label"], "num": row["num"], "origin": row["origin"], "dest": row["dest"],
                    "time": row["time_spoken"], "next": row["next_station"],
                }
                (inbound if row["direction"] == "inbound" else outbound).append(item)
            data = {train["label"]: train for train in inbound + outbound}
            result = (inbound, outbound, data, version)

    if TRAIN_CACHE_SECONDS:
        with _TRAIN_LIST_CACHE_LOCK:
            _TRAIN_LIST_CACHE[cache_key] = (time.monotonic(), _clone_train_result(result))
    return _clone_train_result(result)


def build_train_label(num, hhmm, origin, dest):
    return f"{num} ({hhmm}) {origin} - {dest}"


def save_train_record(form):
    train_id = (form.get("train_id") or "").strip()
    version_id = int(form.get("version_id") or 0)
    direction = (form.get("direction") or "").strip()
    num = (form.get("num") or "").strip()
    origin = (form.get("origin") or "").strip()
    dest = (form.get("dest") or "").strip()
    hhmm = (form.get("time_hhmm") or "").strip()
    next_station = (form.get("next_station") or "").strip()
    pattern = (form.get("service_pattern") or "daily").strip()
    enabled = 1 if form.get("enabled") in {"1", "on", "true"} else 0
    service_dates = normalize_date_list(form.get("service_dates") or "")

    if not version_id or direction not in DIRECTION_LABELS or pattern not in SERVICE_LABELS:
        raise ValueError("ข้อมูลประเภทตารางไม่ถูกต้อง")
    if not all([num, origin, dest, hhmm]):
        raise ValueError("กรุณากรอกเลขขบวน ต้นทาง ปลายทาง และเวลาให้ครบ")
    if not re.fullmatch(r"(?:[01]?\d|2[0-3]):[0-5]\d", hhmm):
        raise ValueError("เวลาต้องเป็นรูปแบบ HH:MM เช่น 05:30")
    hhmm = f"{int(hhmm.split(':')[0]):02d}:{hhmm.split(':')[1]}"
    if pattern in {"holiday", "custom"} and not service_dates:
        raise ValueError("รูปแบบนี้ต้องระบุวันที่ให้บริการอย่างน้อย 1 วัน")

    label = build_train_label(num, hhmm, origin, dest)
    spoken = spoken_time_from_hhmm(hhmm)
    with get_db() as conn:
        duplicate = conn.execute(
            """SELECT id FROM trains WHERE version_id=? AND direction=? AND num=? AND time_hhmm=?
               AND id<>?""",
            (version_id, direction, num, hhmm, int(train_id or 0)),
        ).fetchone()
        if duplicate:
            raise ValueError(f"พบข้อมูลซ้ำ: ขบวน {num} เวลา {hhmm} มีอยู่ในตารางนี้แล้ว")
        if train_id:
            conn.execute(
                """UPDATE trains SET direction=?,num=?,label=?,origin=?,dest=?,time_hhmm=?,time_spoken=?,
                   next_station=?,service_pattern=?,service_dates=?,enabled=?,updated_at=? WHERE id=? AND version_id=?""",
                (direction, num, label, origin, dest, hhmm, spoken, next_station, pattern,
                 service_dates, enabled, now_iso(), int(train_id), version_id),
            )
        else:
            conn.execute(
                """INSERT INTO trains(version_id,direction,num,label,origin,dest,time_hhmm,time_spoken,next_station,
                   service_pattern,service_dates,enabled,created_at,updated_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (version_id, direction, num, label, origin, dest, hhmm, spoken, next_station,
                 pattern, service_dates, enabled, now_iso(), now_iso()),
            )


def parse_bool(value, default=True):
    if value is None or value == "":
        return 1 if default else 0
    return 0 if str(value).strip().lower() in {"0", "false", "no", "off", "ปิด"} else 1


def canonical_import_row(row):
    aliases = {
        "direction": ["direction", "ทิศทาง", "ประเภทขา"],
        "num": ["num", "train_no", "เลขขบวน", "ขบวน"],
        "origin": ["origin", "ต้นทาง"],
        "dest": ["dest", "destination", "ปลายทาง"],
        "time_hhmm": ["time", "time_hhmm", "เวลา"],
        "next_station": ["next", "next_station", "สถานีต่อไป"],
        "service_pattern": ["service_pattern", "วันให้บริการ", "รูปแบบวัน"],
        "service_dates": ["service_dates", "วันที่ให้บริการ", "วันที่ระบุ"],
        "enabled": ["enabled", "active", "เปิดใช้งาน"],
    }
    clean = {str(k).strip(): v for k, v in row.items() if k is not None}
    result = {}
    for key, names in aliases.items():
        result[key] = next((clean.get(name) for name in names if name in clean), "")
    direction_map = {
        "inbound": "inbound", "ขาเข้า": "inbound", "ขาเข้ากรุงเทพ": "inbound", "เข้า": "inbound",
        "outbound": "outbound", "ขาออก": "outbound", "ขาออกตะวันออก": "outbound", "ออก": "outbound",
    }
    pattern_map = {
        "daily": "daily", "ทุกวัน": "daily",
        "weekday": "weekday", "วันธรรมดา": "weekday", "จันทร์-ศุกร์": "weekday",
        "weekend": "weekend", "เสาร์-อาทิตย์": "weekend", "วันเสาร์-อาทิตย์": "weekend",
        "holiday": "holiday", "วันหยุด": "holiday",
        "custom": "custom", "วันที่ระบุ": "custom", "เฉพาะวันที่": "custom",
    }
    result["direction"] = direction_map.get(str(result["direction"]).strip().lower(), str(result["direction"]).strip().lower())
    result["service_pattern"] = pattern_map.get(str(result["service_pattern"]).strip().lower(), str(result["service_pattern"]).strip().lower() or "daily")
    value = result["time_hhmm"]
    if isinstance(value, datetime):
        value = value.strftime("%H:%M")
    elif isinstance(value, dt_time):
        value = value.strftime("%H:%M")
    elif isinstance(value, (int, float)) and 0 <= value < 1:
        total_minutes = round(value * 24 * 60)
        value = f"{(total_minutes // 60) % 24:02d}:{total_minutes % 60:02d}"
    else:
        value = str(value or "").strip()
        match = re.search(r"(\d{1,2}):(\d{2})", value)
        if match:
            value = f"{int(match.group(1)):02d}:{match.group(2)}"
    result["time_hhmm"] = value
    result["enabled"] = parse_bool(result["enabled"])
    return result


def import_schedule_file(file_storage, version_id):
    filename = (file_storage.filename or "").lower()
    rows = []
    if filename.endswith(".csv"):
        raw = file_storage.read()
        text = None
        for encoding in ("utf-8-sig", "utf-8", "cp874"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("อ่านไฟล์ CSV ไม่ได้ กรุณาบันทึกเป็น UTF-8")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("ยังไม่มี openpyxl ให้ติดตั้งด้วย pip install openpyxl") from exc
        workbook = load_workbook(file_storage, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(v or "").strip() for v in next(iterator, [])]
        rows = [dict(zip(headers, values)) for values in iterator if any(v not in (None, "") for v in values)]
    else:
        raise ValueError("รองรับเฉพาะไฟล์ .csv และ .xlsx")

    inserted = 0
    skipped = []
    with get_db() as conn:
        for line_no, raw_row in enumerate(rows, start=2):
            try:
                row = canonical_import_row(raw_row)
                data = {
                    "version_id": str(version_id), "direction": row["direction"], "num": str(row["num"] or "").strip(),
                    "origin": str(row["origin"] or "").strip(), "dest": str(row["dest"] or "").strip(),
                    "time_hhmm": row["time_hhmm"], "next_station": str(row["next_station"] or "").strip(),
                    "service_pattern": row["service_pattern"], "service_dates": str(row["service_dates"] or "").strip(),
                    "enabled": "1" if row["enabled"] else "0",
                }
                # ตรวจข้อมูลซ้ำก่อนเพิ่มทุกแถว
                hhmm = data["time_hhmm"]
                duplicate = conn.execute(
                    "SELECT id FROM trains WHERE version_id=? AND direction=? AND num=? AND time_hhmm=?",
                    (version_id, data["direction"], data["num"], hhmm),
                ).fetchone()
                if duplicate:
                    skipped.append(f"แถว {line_no}: ขบวน {data['num']} เวลา {hhmm} ซ้ำ")
                    continue
                service_dates = normalize_date_list(data["service_dates"])
                if data["direction"] not in DIRECTION_LABELS or data["service_pattern"] not in SERVICE_LABELS:
                    raise ValueError("ทิศทางหรือรูปแบบวันไม่ถูกต้อง")
                if not all([data["num"], data["origin"], data["dest"], hhmm]):
                    raise ValueError("ข้อมูลไม่ครบ")
                if not re.fullmatch(r"(?:[01]?\d|2[0-3]):[0-5]\d", hhmm):
                    raise ValueError("เวลาไม่ใช่ HH:MM")
                hhmm = f"{int(hhmm.split(':')[0]):02d}:{hhmm.split(':')[1]}"
                label = build_train_label(data["num"], hhmm, data["origin"], data["dest"])
                conn.execute(
                    """INSERT INTO trains(version_id,direction,num,label,origin,dest,time_hhmm,time_spoken,next_station,
                       service_pattern,service_dates,enabled,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (version_id, data["direction"], data["num"], label, data["origin"], data["dest"], hhmm,
                     spoken_time_from_hhmm(hhmm), data["next_station"], data["service_pattern"], service_dates,
                     1 if data["enabled"] == "1" else 0, now_iso(), now_iso()),
                )
                inserted += 1
            except Exception as exc:
                skipped.append(f"แถว {line_no}: {exc}")
    return inserted, skipped


def announcement_title(tab_index):
    try:
        tab_index = int(tab_index)
    except (TypeError, ValueError):
        return "ไม่ทราบประเภท"
    item = next((item for item in ANNOUNCEMENT_BUTTONS if item["idx"] == tab_index), None)
    return item["title"] if item else "ไม่ทราบประเภท"


def insert_history(payload, user):
    with get_db() as conn:
        cursor = conn.execute(
            """INSERT INTO announcement_history(started_at,user_id,username,train_num,announcement_type,
               announce_mode,voice,platform,message,pause_times)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (now_iso(), user["id"], user["display_name"], payload.get("num", ""),
             announcement_title(payload.get("tab_index")), payload.get("announce_mode", "thai_only"),
             payload.get("thai_voice", VOICE_NAME), payload.get("platform", ""), "", "[]"),
        )
        history_id = cursor.lastrowid
        conn.execute(
            "INSERT INTO announcement_events(history_id,event_type,event_at,details) VALUES(?,?,?,?)",
            (history_id, "start", now_iso(), json.dumps(payload, ensure_ascii=False)),
        )
    return history_id


def add_history_event(history_id, event_type, details=None):
    event_at = now_iso()
    details = details or {}
    with get_db() as conn:
        row = conn.execute("SELECT * FROM announcement_history WHERE id=?", (history_id,)).fetchone()
        if not row:
            raise ValueError("ไม่พบประวัติรายการนี้")
        conn.execute(
            "INSERT INTO announcement_events(history_id,event_type,event_at,details) VALUES(?,?,?,?)",
            (history_id, event_type, event_at, json.dumps(details, ensure_ascii=False)),
        )
        if event_type == "generated":
            conn.execute(
                "UPDATE announcement_history SET message=?,generation_ms=? WHERE id=?",
                (details.get("message", ""), details.get("generation_ms"), history_id),
            )
        elif event_type == "pause":
            pauses = json.loads(row["pause_times"] or "[]")
            pauses.append(event_at)
            conn.execute("UPDATE announcement_history SET pause_times=? WHERE id=?", (json.dumps(pauses), history_id))
        elif event_type == "stop":
            conn.execute(
                "UPDATE announcement_history SET stop_time=?,playback_success=0,completed_at=? WHERE id=?",
                (event_at, event_at, history_id),
            )
        elif event_type == "success":
            conn.execute(
                "UPDATE announcement_history SET playback_success=1,completed_at=?,failure_reason=NULL WHERE id=?",
                (event_at, history_id),
            )
        elif event_type == "failed":
            conn.execute(
                "UPDATE announcement_history SET playback_success=0,completed_at=?,failure_reason=? WHERE id=?",
                (event_at, str(details.get("reason", ""))[:1000], history_id),
            )
    return event_at


_HEALTH_CACHE = {"at": 0.0, "tts": None}


def backend_health_checks():
    chime = BASE_DIR / CHIME_FILENAME
    disk = shutil.disk_usage(BASE_DIR)
    used_percent = round((disk.used / disk.total) * 100, 1) if disk.total else 0
    audio_writable = os.access(AUDIO_DIR, os.W_OK)
    tts_available = shutil.which("edge-tts") is not None
    tts_connected = False
    tts_detail = "ไม่พบโปรแกรม edge-tts"
    now_ts = time.time()
    if tts_available:
        if now_ts - _HEALTH_CACHE["at"] > 60 or _HEALTH_CACHE["tts"] is None:
            try:
                result = subprocess.run(
                    ["edge-tts", "--list-voices"], capture_output=True, text=True,
                    timeout=8, encoding="utf-8", errors="replace",
                )
                _HEALTH_CACHE["tts"] = (result.returncode == 0, (result.stderr or "").strip())
            except Exception as exc:
                _HEALTH_CACHE["tts"] = (False, str(exc))
            _HEALTH_CACHE["at"] = now_ts
        tts_connected, error = _HEALTH_CACHE["tts"]
        tts_detail = "เชื่อมต่อและเรียกรายการเสียงได้" if tts_connected else (error or "เชื่อมต่อ TTS ไม่สำเร็จ")

    database_ok = False
    database_detail = ""
    try:
        with get_db() as conn:
            conn.execute("SELECT 1").fetchone()
        database_ok = True
        database_detail = (f"Supabase PostgreSQL พร้อมใช้ · connection pool {POSTGRES_POOL_MIN_SIZE}–{POSTGRES_POOL_MAX_SIZE}" if USE_POSTGRES else f"{DB_PATH.name} พร้อมใช้งาน")
    except Exception as exc:
        database_detail = f"เชื่อมต่อฐานข้อมูลไม่สำเร็จ: {str(exc)[:180]}"

    checks = [
        {"key": "backend", "label": "Backend", "ok": True, "detail": "Flask ตอบสนองตามปกติ"},
        {"key": "chime", "label": "ไฟล์เสียงเตือน", "ok": chime.exists() and chime.stat().st_size > 0,
         "detail": "พบ chime.mp3" if chime.exists() else "ไม่พบ chime.mp3 ในโฟลเดอร์โปรแกรม"},
        {"key": "tts", "label": "การเชื่อมต่อ TTS", "ok": tts_available and tts_connected, "detail": tts_detail},
        {"key": "storage", "label": "พื้นที่จัดเก็บ", "ok": used_percent < 85,
         "detail": f"ใช้งาน {used_percent}% · เหลือ {round(disk.free / (1024**3), 2)} GB"},
        {"key": "audio_dir", "label": "โฟลเดอร์ไฟล์เสียง", "ok": AUDIO_DIR.exists() and audio_writable,
         "detail": f"พร้อมเขียนไฟล์ · มีไฟล์เสียง {len(list(AUDIO_DIR.glob('*.mp3')))} ไฟล์"},
        {"key": "database", "label": "ฐานข้อมูล", "ok": database_ok,
         "detail": database_detail},
    ]
    return checks


init_database()

HTML_PAGE = r"""
<!DOCTYPE html>
<html lang="th">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ระบบประกาศสถานีคลองบางพระ</title>
    <link rel="preload" href="/audio/chime.mp3" as="audio" type="audio/mpeg">
    <style>
        :root {
            --maroon: #800000;
            --maroon-dark: #5b0000;
            --gold: #c7a12a;
            --cream: #fffaf1;
            --paper: #ffffff;
            --ink: #251d1d;
            --muted: #716464;
            --line: #e9dece;
            --green: #177744;
            --red: #b42318;
            --shadow: 0 12px 34px rgba(83, 28, 28, .10);
        }
        * { box-sizing: border-box; }
        body {
            margin: 0;
            min-height: 100vh;
            padding: 16px;
            color: var(--ink);
            font-family: "Sarabun", "Noto Sans Thai", "Segoe UI", sans-serif;
            background: linear-gradient(145deg, #fffaf1, #f4eadc);
        }
        button, input, select, textarea { font: inherit; }
        button { cursor: pointer; }
        .app { max-width: 1120px; margin: 0 auto; }
        .system-nav {
            display: flex; flex-wrap: wrap; align-items: center; gap: 8px;
            margin-top: 11px; padding: 9px;
            border: 1px solid var(--line); border-radius: 16px;
            background: rgba(255,255,255,.92); box-shadow: var(--shadow);
        }
        .system-nav a {
            padding: 9px 12px; border-radius: 11px; color: var(--maroon-dark);
            text-decoration: none; font-weight: 850; background: #fff7ed;
        }
        .system-nav a.active, .system-nav a:hover { color: white; background: var(--maroon); }
        .nav-user { margin-left: auto; color: var(--muted); font-size: 12px; font-weight: 750; }
        .topbar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 16px;
            padding: 18px 20px;
            color: white;
            border-radius: 22px;
            background: linear-gradient(135deg, var(--maroon-dark), var(--maroon));
            box-shadow: var(--shadow);
        }
        .brand { display: flex; align-items: center; gap: 13px; }
        .logo {
            width: 52px; height: 52px; flex: 0 0 auto;
            display: grid; place-items: center;
            border-radius: 16px;
            font-size: 27px;
            color: var(--maroon);
            background: linear-gradient(135deg, #fff4c4, var(--gold));
        }
        h1 { margin: 0; font-size: clamp(21px, 3vw, 30px); line-height: 1.2; }
        .subtitle { margin: 4px 0 0; opacity: .88; font-size: 13px; }
        .status {
            min-width: 180px;
            padding: 10px 14px;
            border: 1px solid rgba(255,255,255,.25);
            border-radius: 999px;
            text-align: center;
            font-weight: 800;
            background: rgba(255,255,255,.12);
        }
        .progress { display: none; height: 4px; margin-top: 10px; overflow: hidden; border-radius: 99px; background: rgba(255,255,255,.2); }
        .progress.active { display: block; }
        .progress::after {
            content: ""; display: block; width: 35%; height: 100%;
            background: linear-gradient(90deg, transparent, #fff5a5, transparent);
            animation: loading 1s linear infinite;
        }
        @keyframes loading { from { transform: translateX(-100%); } to { transform: translateX(300%); } }

        .layout {
            display: grid;
            grid-template-columns: minmax(0, 1fr) minmax(320px, .74fr);
            gap: 16px;
            align-items: start;
            margin-top: 16px;
        }
        .stack { display: grid; gap: 16px; }
        .card {
            background: rgba(255,255,255,.96);
            border: 1px solid var(--line);
            border-radius: 20px;
            box-shadow: var(--shadow);
            overflow: hidden;
        }
        .card-head { padding: 15px 18px; border-bottom: 1px solid var(--line); background: #fffdf9; }
        .step-title { display: flex; align-items: center; gap: 10px; margin: 0; color: var(--maroon-dark); font-size: 18px; }
        .step {
            width: 30px; height: 30px; display: grid; place-items: center;
            border-radius: 10px; color: white; background: var(--maroon); font-size: 14px;
        }
        .card-body { padding: 17px; }
        .helper { margin: 6px 0 0; color: var(--muted); font-size: 12.5px; line-height: 1.45; }
        label { display: block; margin: 0 0 6px; font-weight: 800; }
        input, select, textarea {
            width: 100%; padding: 12px 13px;
            border: 1px solid #d9ccb9; border-radius: 13px;
            background: white; color: var(--ink); outline: none;
        }
        input:focus, select:focus, textarea:focus { border-color: var(--gold); box-shadow: 0 0 0 4px rgba(199,161,42,.14); }
        textarea { min-height: 92px; resize: vertical; }
        .field-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 11px; }
        .full { grid-column: 1 / -1; }

        .language-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 9px; }
        .lang-btn {
            min-height: 62px; padding: 10px;
            border: 1px solid #dfd1bd; border-radius: 15px;
            background: #fff; color: var(--ink); font-weight: 850;
        }
        .lang-btn small { display: block; margin-top: 2px; color: var(--muted); font-weight: 600; }
        .lang-btn.active { border-color: var(--maroon); color: var(--maroon); background: #fff1f1; box-shadow: inset 0 0 0 1px var(--maroon); }

        .train-summary {
            display: grid; grid-template-columns: auto 1fr; gap: 10px 12px;
            margin-top: 12px; padding: 14px;
            border-radius: 16px; border: 1px solid #eadcc5; background: var(--cream);
        }
        .train-number {
            grid-row: 1 / span 2; align-self: stretch;
            min-width: 74px; display: grid; place-items: center;
            border-radius: 13px; background: var(--maroon); color: white;
            font-size: 22px; font-weight: 900;
        }
        .route { font-size: 17px; font-weight: 900; color: var(--maroon-dark); }
        .train-meta { color: var(--muted); font-size: 13px; line-height: 1.5; }
        .platform-row { display: grid; grid-template-columns: 1fr 140px; gap: 11px; margin-top: 12px; }
        .train-pickers { display: grid; gap: 14px; }
        .train-picker {
            padding: 14px; border: 1px solid #e7d9c6; border-radius: 18px;
            background: linear-gradient(145deg, #fff, #fffaf2);
        }
        .train-picker.primary-train { border-color: rgba(128,0,0,.38); box-shadow: inset 4px 0 0 var(--maroon); }
        .train-picker-head {
            display: flex; align-items: center; justify-content: space-between;
            gap: 10px; margin-bottom: 10px;
        }
        .train-picker-title { display: flex; align-items: center; gap: 9px; font-weight: 900; color: var(--maroon-dark); }
        .train-order {
            width: 30px; height: 30px; display: grid; place-items: center;
            border-radius: 10px; color: white; background: var(--maroon); font-weight: 900;
        }
        .train-role {
            padding: 5px 9px; border-radius: 999px; background: #f5ecdf;
            color: var(--muted); font-size: 11px; font-weight: 800;
        }
        .train-picker .train-summary { margin-top: 10px; }
        .train-picker .platform-row { grid-template-columns: minmax(0,1fr); }
        .station-row { margin-top: 13px; }

        .announce-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 9px; }
        .announce-option {
            min-height: 74px; padding: 12px;
            border: 1px solid #e5d7c4; border-radius: 15px;
            text-align: left; color: var(--ink); background: linear-gradient(145deg, #fff, #fff9ef);
        }
        .announce-option strong { display: block; color: var(--maroon-dark); font-size: 15px; }
        .announce-option span { display: block; margin-top: 3px; color: var(--muted); font-size: 12px; line-height: 1.35; }
        .announce-option.active { border-color: var(--maroon); background: #fff0f0; box-shadow: inset 0 0 0 1px var(--maroon); }
        .announce-option:disabled { opacity: .55; cursor: not-allowed; }

        .conditional {
            display: none; margin-top: 13px; padding: 14px;
            border-radius: 16px; border: 1px dashed rgba(128,0,0,.35); background: #fffaf3;
        }
        .conditional.show { display: block; }
        .conditional-title { margin: 0 0 10px; color: var(--maroon); font-weight: 900; }
        .advanced { margin-top: 12px; }
        .advanced summary { cursor: pointer; color: var(--maroon); font-weight: 850; }
        .advanced-content { margin-top: 12px; }

        .sticky { position: sticky; top: 16px; }
        .selected-type {
            padding: 12px 14px; margin-bottom: 12px;
            border-radius: 14px; background: #f7f1e8; color: var(--muted);
        }
        .selected-type b { color: var(--maroon-dark); }
        .preview {
            min-height: 150px; max-height: 390px; overflow: auto;
            padding: 15px; border: 1px solid #eadcc5; border-radius: 16px;
            background: #fffaf0; line-height: 1.7; font-size: 14px;
        }
        .action-stack { display: grid; gap: 9px; margin-top: 13px; }
        .primary, .secondary, .danger, .pause-btn {
            width: 100%; border: 0; border-radius: 14px; padding: 14px;
            color: white; font-weight: 900;
        }
        .primary { background: linear-gradient(135deg, var(--maroon-dark), var(--maroon)); font-size: 16px; }
        .pause-btn { background: #b06b00; }
        .secondary { background: #665b55; }
        .danger { background: var(--red); }
        .primary:disabled, .pause-btn:disabled, .danger:disabled { opacity: .45; cursor: not-allowed; }
        .playback-controls { display: grid; grid-template-columns: 1.25fr 1fr 1fr; gap: 9px; }
        .mini-note { margin-top: 12px; color: var(--muted); font-size: 12px; line-height: 1.5; }
        .voice-quick-panel {
            margin-top: 14px;
            padding: 16px;
            border: 2px solid rgba(128,0,0,.24);
            border-radius: 20px;
            background: linear-gradient(145deg, #ffffff, #fff7e9);
            box-shadow: var(--shadow);
        }
        .voice-quick-head {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 12px;
            margin-bottom: 12px;
        }
        .voice-quick-title {
            margin: 0;
            color: var(--maroon-dark);
            font-size: 18px;
            font-weight: 900;
        }
        .voice-quick-note {
            color: var(--muted);
            font-size: 12.5px;
        }
        .voice-choice-grid {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr)) auto;
            gap: 10px;
        }
        .voice-choice-btn, .voice-test-btn {
            min-height: 58px;
            padding: 11px 14px;
            border: 1px solid #ddcfba;
            border-radius: 15px;
            background: white;
            color: var(--ink);
            font-weight: 900;
            text-align: left;
        }
        .voice-choice-btn small {
            display: block;
            margin-top: 3px;
            color: var(--muted);
            font-weight: 600;
        }
        .voice-choice-btn.active {
            border-color: var(--maroon);
            color: var(--maroon-dark);
            background: #fff0f0;
            box-shadow: inset 0 0 0 1px var(--maroon);
        }
        .voice-test-btn {
            min-width: 165px;
            text-align: center;
            color: white;
            border-color: var(--maroon);
            background: var(--maroon);
        }
        .voice-choice-btn:disabled, .voice-test-btn:disabled {
            opacity: .45;
            cursor: not-allowed;
        }
        .hidden { display: none !important; }

        @media (max-width: 860px) {
            body { padding: 9px; }
            .topbar { border-radius: 18px; padding: 15px; }
            .status { min-width: auto; }
            .layout { grid-template-columns: 1fr; }
            .sticky { position: static; }
        }
        @media (max-width: 560px) {
            .topbar { align-items: flex-start; }
            .logo { width: 44px; height: 44px; font-size: 23px; }
            .status { font-size: 12px; padding: 8px 10px; }
            .language-grid { grid-template-columns: 1fr; }
            .lang-btn { min-height: 52px; text-align: left; }
            .announce-grid, .field-grid, .platform-row { grid-template-columns: 1fr; }
            .train-summary { grid-template-columns: 64px 1fr; }
            .train-number { min-width: 64px; font-size: 19px; }
            .voice-choice-grid { grid-template-columns: 1fr; }
            .voice-test-btn { width: 100%; min-width: 0; }
            .playback-controls { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>
<main class="app">
    <header class="topbar">
        <div class="brand">
            <div class="logo">🚆</div>
            <div>
                <h1>ระบบประกาศสถานีคลองบางพระ</h1>
                <p class="subtitle">เลือกภาษา เลือกเสียง แล้วกดประกาศได้ทันที</p>
            </div>
        </div>
        <div class="status" id="statusText">พร้อมใช้งาน</div>
        <div class="progress" id="loadingBar"></div>
    </header>

    <nav class="system-nav">
        <a class="active" href="{{ url_for('index') }}">📢 หน้าประกาศ</a>
        {% if current_user.role == 'admin' %}
        <a href="{{ url_for('admin_schedules') }}">🚆 จัดการตารางรถ</a>
        <a href="{{ url_for('admin_users') }}">👥 บัญชีผู้ใช้</a>
        {% endif %}
        {% if current_user.role in ['admin', 'auditor'] %}
        <a href="{{ url_for('history_page') }}">🕘 ประวัติการประกาศ</a>
        {% endif %}
        <a href="{{ url_for('health_page') }}">🩺 ตรวจสุขภาพระบบ</a>
        <span class="nav-user">{{ current_user.display_name }} · {{ role_labels[current_user.role] }}</span>
        <a href="{{ url_for('logout') }}">ออกจากระบบ</a>
    </nav>

    <section class="voice-quick-panel" id="thaiVoiceSettings">
        <div class="voice-quick-head">
            <div>
                <h2 class="voice-quick-title">🔊 เลือกเสียงประกาศ</h2>
                <div class="voice-quick-note" id="voiceHelper">เสียงที่เลือกจะใช้เพศเดียวกันทั้งภาษาไทยและภาษาอังกฤษ</div>
            </div>
        </div>
        <input type="hidden" id="thai_voice" value="{{ voice_name }}">
        <div class="voice-choice-grid">
            <button type="button"
                    class="voice-choice-btn {% if voice_name == 'th-TH-PremwadeeNeural' %}active{% endif %}"
                    data-voice="th-TH-PremwadeeNeural"
                    onclick="setThaiVoice('th-TH-PremwadeeNeural', this)">
                👩 เสียงผู้หญิง
                <small>ไทย: Premwadee · อังกฤษ: Jenny</small>
            </button>
            <button type="button"
                    class="voice-choice-btn {% if voice_name == 'th-TH-NiwatNeural' %}active{% endif %}"
                    data-voice="th-TH-NiwatNeural"
                    onclick="setThaiVoice('th-TH-NiwatNeural', this)">
                👨 เสียงผู้ชาย
                <small>ไทย: Niwat · อังกฤษ: Guy</small>
            </button>
            <button type="button" class="voice-test-btn" onclick="testStationVoice()">▶ ทดลองเสียง</button>
        </div>
    </section>

    <section class="layout">
        <div class="stack">
            <section class="card">
                <div class="card-head"><h2 class="step-title"><span class="step">1</span> เลือกภาษาประกาศ</h2></div>
                <div class="card-body">
                    <input type="hidden" id="announce_mode" value="thai_only">
                    <div class="language-grid">
                        <button type="button" class="lang-btn active" data-mode="thai_only" onclick="setLanguage('thai_only', this)">🇹🇭 ภาษาไทย<small>ประกาศภาษาเดียว</small></button>
                        <button type="button" class="lang-btn" data-mode="english_only" onclick="setLanguage('english_only', this)">🇬🇧 English<small>English only</small></button>
                        <button type="button" class="lang-btn" data-mode="bilingual" onclick="setLanguage('bilingual', this)">🇹🇭 + 🇬🇧 สองภาษา<small>ไทย แล้วอังกฤษ</small></button>
                    </div>
                    <div class="helper">เสียงภาษาอังกฤษจะเปลี่ยนเป็นผู้หญิงหรือผู้ชายให้ตรงกับเสียงที่เลือกด้านบนโดยอัตโนมัติ</div>
                </div>
            </section>

            <section class="card">
                <div class="card-head"><h2 class="step-title"><span class="step">2</span> เลือกขบวนและชานชาลา</h2></div>
                <div class="card-body">
                    <p class="helper" style="margin:0 0 12px;">เลือกได้สูงสุด 3 ขบวน โดยขบวนที่ 1 ใช้กับประกาศทั่วไป ส่วนปุ่ม “รถเข้าพร้อมกัน 2–3 ขบวน” จะนำขบวนที่เลือกทั้งหมดมาประกาศร่วมกัน</p>

                    <div class="train-pickers">
                        <div class="train-picker primary-train">
                            <div class="train-picker-head">
                                <div class="train-picker-title"><span class="train-order">1</span> ขบวนที่ 1</div>
                                <span class="train-role">ขบวนหลัก</span>
                            </div>
                            <label for="train_select">เลือกขบวนรถ</label>
                            <select id="train_select" onchange="autoFill(1)">
                                <option value="">-- เลือกขบวนรถ --</option>
                                <optgroup label="ขาเข้า กรุงเทพ (หัวลำโพง)">
                                {% for train in inbound %}<option value="{{ train.label }}">{{ train.label }}</option>{% endfor %}
                                </optgroup>
                                <optgroup label="ขาออก ไปทางตะวันออก">
                                {% for train in outbound %}<option value="{{ train.label }}">{{ train.label }}</option>{% endfor %}
                                </optgroup>
                            </select>
                            <div class="train-summary" id="trainSummary">
                                <div class="train-number" id="summaryNum">–</div>
                                <div class="route" id="summaryRoute">ยังไม่ได้เลือกขบวนรถ</div>
                                <div class="train-meta" id="summaryMeta">เมื่อเลือกขบวน ระบบจะเติมข้อมูลให้อัตโนมัติ</div>
                            </div>
                            <div class="platform-row">
                                <div>
                                    <label for="platform">ชานชาลาที่</label>
                                    <select id="platform" onchange="syncPlatformDefaults(1)">
                                        <option value="1" selected>ชานชาลาที่ 1</option>
                                        <option value="2">ชานชาลาที่ 2</option>
                                        <option value="3">ชานชาลาที่ 3</option>
                                    </select>
                                </div>
                            </div>
                            <details class="advanced">
                                <summary>แก้ไขรายละเอียดขบวนที่ 1</summary>
                                <div class="advanced-content field-grid">
                                    <div><label>ขบวนที่</label><input type="text" id="num" oninput="refreshSummary(1)"></div>
                                    <div><label>เวลา</label><input type="text" id="time" oninput="refreshSummary(1)"></div>
                                    <div><label>ต้นทาง</label><input type="text" id="origin" oninput="refreshSummary(1)"></div>
                                    <div><label>ปลายทาง</label><input type="text" id="dest" oninput="refreshSummary(1)"></div>
                                    <div class="full"><label>สถานีต่อไป</label><input type="text" id="next_station" oninput="refreshSummary(1)"></div>
                                </div>
                            </details>
                        </div>

                        <div class="train-picker">
                            <div class="train-picker-head">
                                <div class="train-picker-title"><span class="train-order">2</span> ขบวนที่ 2</div>
                                <span class="train-role">ไม่บังคับ</span>
                            </div>
                            <label for="train_select_2">เลือกขบวนรถ</label>
                            <select id="train_select_2" onchange="autoFill(2)">
                                <option value="">-- ยังไม่เลือกขบวนที่ 2 --</option>
                                <optgroup label="ขาเข้า กรุงเทพ (หัวลำโพง)">
                                {% for train in inbound %}<option value="{{ train.label }}">{{ train.label }}</option>{% endfor %}
                                </optgroup>
                                <optgroup label="ขาออก ไปทางตะวันออก">
                                {% for train in outbound %}<option value="{{ train.label }}">{{ train.label }}</option>{% endfor %}
                                </optgroup>
                            </select>
                            <div class="train-summary" id="trainSummary2">
                                <div class="train-number" id="summaryNum2">–</div>
                                <div class="route" id="summaryRoute2">ยังไม่ได้เลือกขบวนที่ 2</div>
                                <div class="train-meta" id="summaryMeta2">ใช้เมื่อมีขบวนรถเข้าพร้อมกัน</div>
                            </div>
                            <div class="platform-row">
                                <div>
                                    <label for="platform_2">ชานชาลาที่</label>
                                    <select id="platform_2" onchange="syncPlatformDefaults(2)">
                                        <option value="1">ชานชาลาที่ 1</option>
                                        <option value="2" selected>ชานชาลาที่ 2</option>
                                        <option value="3">ชานชาลาที่ 3</option>
                                    </select>
                                </div>
                            </div>
                            <details class="advanced">
                                <summary>แก้ไขรายละเอียดขบวนที่ 2</summary>
                                <div class="advanced-content field-grid">
                                    <div><label>ขบวนที่</label><input type="text" id="num_2" oninput="refreshSummary(2)"></div>
                                    <div><label>เวลา</label><input type="text" id="time_2" oninput="refreshSummary(2)"></div>
                                    <div><label>ต้นทาง</label><input type="text" id="origin_2" oninput="refreshSummary(2)"></div>
                                    <div><label>ปลายทาง</label><input type="text" id="dest_2" oninput="refreshSummary(2)"></div>
                                    <div class="full"><label>สถานีต่อไป</label><input type="text" id="next_station_2" oninput="refreshSummary(2)"></div>
                                </div>
                            </details>
                        </div>

                        <div class="train-picker">
                            <div class="train-picker-head">
                                <div class="train-picker-title"><span class="train-order">3</span> ขบวนที่ 3</div>
                                <span class="train-role">ไม่บังคับ</span>
                            </div>
                            <label for="train_select_3">เลือกขบวนรถ</label>
                            <select id="train_select_3" onchange="autoFill(3)">
                                <option value="">-- ยังไม่เลือกขบวนที่ 3 --</option>
                                <optgroup label="ขาเข้า กรุงเทพ (หัวลำโพง)">
                                {% for train in inbound %}<option value="{{ train.label }}">{{ train.label }}</option>{% endfor %}
                                </optgroup>
                                <optgroup label="ขาออก ไปทางตะวันออก">
                                {% for train in outbound %}<option value="{{ train.label }}">{{ train.label }}</option>{% endfor %}
                                </optgroup>
                            </select>
                            <div class="train-summary" id="trainSummary3">
                                <div class="train-number" id="summaryNum3">–</div>
                                <div class="route" id="summaryRoute3">ยังไม่ได้เลือกขบวนที่ 3</div>
                                <div class="train-meta" id="summaryMeta3">ใช้เมื่อมีขบวนรถเข้าพร้อมกัน 3 ขบวน</div>
                            </div>
                            <div class="platform-row">
                                <div>
                                    <label for="platform_3">ชานชาลาที่</label>
                                    <select id="platform_3" onchange="syncPlatformDefaults(3)">
                                        <option value="1">ชานชาลาที่ 1</option>
                                        <option value="2">ชานชาลาที่ 2</option>
                                        <option value="3" selected>ชานชาลาที่ 3</option>
                                    </select>
                                </div>
                            </div>
                            <details class="advanced">
                                <summary>แก้ไขรายละเอียดขบวนที่ 3</summary>
                                <div class="advanced-content field-grid">
                                    <div><label>ขบวนที่</label><input type="text" id="num_3" oninput="refreshSummary(3)"></div>
                                    <div><label>เวลา</label><input type="text" id="time_3" oninput="refreshSummary(3)"></div>
                                    <div><label>ต้นทาง</label><input type="text" id="origin_3" oninput="refreshSummary(3)"></div>
                                    <div><label>ปลายทาง</label><input type="text" id="dest_3" oninput="refreshSummary(3)"></div>
                                    <div class="full"><label>สถานีต่อไป</label><input type="text" id="next_station_3" oninput="refreshSummary(3)"></div>
                                </div>
                            </details>
                        </div>
                    </div>

                    <div class="station-row">
                        <label for="current">สถานีปัจจุบัน</label>
                        <input type="text" id="current" value="คลองบางพระ">
                    </div>
                </div>
            </section>

            <section class="card">
                <div class="card-head"><h2 class="step-title"><span class="step">3</span> เลือกประเภทประกาศ</h2></div>
                <div class="card-body">
                    <div class="announce-grid">
                        {% for group, buttons in grouped_buttons.items() %}
                            {% for button in buttons %}
                            <button type="button" class="announce-option" data-index="{{ button.idx }}" data-title="{{ button.title }}" onclick="selectAnnouncement({{ button.idx }}, this)">
                                <strong>{{ button.title }}</strong><span>{{ button.hint }}</span>
                            </button>
                            {% endfor %}
                        {% endfor %}
                    </div>

                    <div class="conditional" id="delayFields">
                        <p class="conditional-title">ข้อมูลรถล่าช้า</p>
                        <label for="delay_time">คาดว่าจะถึงเวลา</label>
                        <input type="text" id="delay_time" placeholder="เช่น 19 นาฬิกา 30 นาที">
                    </div>

                    <div class="conditional" id="passFields">
                        <p class="conditional-title">ข้อมูลรถวิ่งผ่าน</p>
                        <div class="field-grid">
                            <div id="trainTypeWrap">
                                <label for="train_type">ประเภทรถ</label>
                                <select id="train_type">
                                    <option value="สินค้า">สินค้า</option>
                                    <option value="ด่วนพิเศษ">ด่วนพิเศษ</option>
                                    <option value="พิเศษ">พิเศษ</option>
                                    <option value="รถจักรเปล่า">รถจักรเปล่า</option>
                                </select>
                            </div>
                            <div>
                                <label for="pass_platform">ชานชาลาที่รถผ่าน</label>
                                <select id="pass_platform">
                                    <option value="1" selected>ชานชาลาที่ 1</option>
                                    <option value="2">ชานชาลาที่ 2</option>
                                    <option value="3">ชานชาลาที่ 3</option>
                                </select>
                            </div>
                        </div>
                    </div>

                    <div class="conditional" id="customFields">
                        <p class="conditional-title">ข้อความประกาศเอง</p>
                        <div id="thaiCustomWrap">
                            <label for="custom_text">ข้อความภาษาไทย</label>
                            <textarea id="custom_text" placeholder="พิมพ์ข้อความภาษาไทยที่ต้องการประกาศ"></textarea>
                        </div>
                        <div id="englishCustomWrap" class="hidden" style="margin-top:10px;">
                            <label for="custom_text_en">English announcement</label>
                            <textarea id="custom_text_en" placeholder="Type the English announcement"></textarea>
                        </div>
                    </div>

                </div>
            </section>
        </div>

        <aside class="card sticky">
            <div class="card-head"><h2 class="step-title"><span class="step">4</span> ตรวจสอบและประกาศ</h2></div>
            <div class="card-body">
                <div class="selected-type" id="selectedType"><b>ยังไม่ได้เลือกประเภทประกาศ</b><br>เลือกปุ่มในขั้นตอนที่ 3 ก่อน</div>
                <div class="preview" id="previewBox"><b>ตัวอย่างข้อความประกาศ</b><br><br>เมื่อกดเริ่มประกาศ ระบบจะสร้างข้อความและไฟล์เสียงตามภาษาที่เลือก</div>
                <div class="action-stack">
                    <div class="playback-controls">
                        <button type="button" class="primary" id="playButton" onclick="playOrResumeAudio()" disabled>▶ เริ่มประกาศ</button>
                        <button type="button" class="pause-btn" id="pauseButton" onclick="pauseAudio()" disabled>⏸ พักเสียง</button>
                        <button type="button" class="danger" id="stopButton" onclick="stopAudio()" disabled>■ หยุดเสียง</button>
                    </div>
                    <button type="button" class="secondary" onclick="clearData()">ล้างข้อมูล</button>
                </div>
                <p class="mini-note">เสียงเตือนจะเล่นก่อนเสียงประกาศ โดยเสียงภาษาไทยและภาษาอังกฤษจะใช้เพศเดียวกันตามปุ่มที่เลือกด้านบน</p>
            </div>
        </aside>
    </section>
</main>

<script>
    const trainData = {{ trains_json | safe }};
    let selectedAnnouncement = null;
    let mainPlayer = null;
    let isBusy = false;
    let sharedAudioContext = null;
    let mobileAudioUnlocked = false;
    let preparedAudioKey = "";
    let preparedAudioPromise = null;
    let preparedAudioData = null;
    let prepareTimer = null;
    let playbackState = "idle"; // idle | loading | playing | paused
    let playbackRunId = 0;
    let activePlaybackCancel = null;
    let activeHistoryId = null;
    let activeHistoryPromise = null;

    function byId(id) { return document.getElementById(id); }
    function value(id) { return (byId(id)?.value || "").trim(); }

    function startHistoryRecord(tabIndex) {
        const payload = collectPayload(tabIndex);
        return fetch("/api/history/start", {
            method: "POST",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify(payload)
        }).then(async response => {
            const data = await response.json();
            if (!response.ok || data.status !== "success") throw new Error(data.message || "บันทึกประวัติไม่สำเร็จ");
            return data.history_id;
        }).catch(error => {
            console.warn("History start failed:", error);
            return null;
        });
    }

    function logHistoryEvent(eventType, details = {}) {
        const source = activeHistoryId ? Promise.resolve(activeHistoryId) : activeHistoryPromise;
        if (!source) return Promise.resolve();
        return source.then(historyId => {
            if (!historyId) return;
            return fetch("/api/history/event", {
                method: "POST",
                headers: { "Content-Type": "application/json" },
                body: JSON.stringify({ history_id: historyId, event_type: eventType, details })
            });
        }).catch(error => console.warn("History event failed:", error));
    }

    function setLanguage(mode, button) {
        byId("announce_mode").value = mode;
        document.querySelectorAll(".lang-btn").forEach(btn => btn.classList.remove("active"));
        if (button) button.classList.add("active");
        updateCustomLanguageFields();

        document.querySelectorAll(".voice-choice-btn, .voice-test-btn").forEach(btn => {
            btn.disabled = false;
        });

        if (mode === "english_only") {
            byId("voiceHelper").innerText = "เสียงที่เลือกจะใช้กับภาษาอังกฤษ: ผู้หญิง = Jenny, ผู้ชาย = Guy";
        } else if (mode === "bilingual") {
            byId("voiceHelper").innerText = "โหมดสองภาษา: ไทยและอังกฤษจะใช้เพศเดียวกันตามเสียงที่เลือก";
        } else {
            byId("voiceHelper").innerText = "เสียงที่เลือกจะใช้กับภาษาไทย และคำลงท้ายจะเปลี่ยนเป็นครับหรือค่ะอัตโนมัติ";
        }
        invalidatePreparedAudio();
        schedulePrepareAnnouncement();
    }

    function setThaiVoice(voice, button) {
        byId("thai_voice").value = voice;
        document.querySelectorAll(".voice-choice-btn").forEach(btn => btn.classList.remove("active"));
        if (button) button.classList.add("active");
        const voiceName = voice === "th-TH-NiwatNeural" ? "เสียงผู้ชาย" : "เสียงผู้หญิง";
        setStatus("เลือก " + voiceName + " แล้ว", "ok");
        invalidatePreparedAudio();
        schedulePrepareAnnouncement();
    }

    function updateCustomLanguageFields() {
        const mode = value("announce_mode") || "thai_only";
        byId("thaiCustomWrap").classList.toggle("hidden", mode === "english_only");
        byId("englishCustomWrap").classList.toggle("hidden", mode === "thai_only");
    }

    function trainSuffix(type) { return type === 1 ? "" : `_${type}`; }
    function summarySuffix(type) { return type === 1 ? "" : String(type); }

    function autoFill(type) {
        const suffix = trainSuffix(type);
        const selectId = type === 1 ? "train_select" : `train_select_${type}`;
        const data = trainData[value(selectId)];
        if (!data) {
            ["num", "origin", "dest", "time", "next_station"].forEach(field => {
                const el = byId(field + suffix);
                if (el) el.value = "";
            });
            refreshSummary(type);
            invalidatePreparedAudio();
            schedulePrepareAnnouncement();
            return;
        }
        byId("num" + suffix).value = data.num || "";
        byId("origin" + suffix).value = data.origin || "";
        byId("dest" + suffix).value = data.dest || "";
        byId("time" + suffix).value = data.time || "";
        byId("next_station" + suffix).value = data.next || "";
        refreshSummary(type);
        invalidatePreparedAudio();
        schedulePrepareAnnouncement();
    }

    function refreshSummary(type = 1) {
        const suffix = trainSuffix(type);
        const summary = summarySuffix(type);
        const num = value("num" + suffix) || "–";
        const origin = value("origin" + suffix);
        const dest = value("dest" + suffix);
        const time = value("time" + suffix);
        const next = value("next_station" + suffix);
        byId("summaryNum" + summary).textContent = num;
        byId("summaryRoute" + summary).textContent = origin && dest
            ? `${origin} → ${dest}`
            : (type === 1 ? "ยังไม่ได้เลือกขบวนรถ" : `ยังไม่ได้เลือกขบวนที่ ${type}`);
        const details = [];
        if (time) details.push(`เวลา ${time}`);
        if (next) details.push(`สถานีต่อไป: ${next}`);
        const emptyText = type === 1
            ? "เมื่อเลือกขบวน ระบบจะเติมข้อมูลให้อัตโนมัติ"
            : (type === 2 ? "ใช้เมื่อมีขบวนรถเข้าพร้อมกัน" : "ใช้เมื่อมีขบวนรถเข้าพร้อมกัน 3 ขบวน");
        byId("summaryMeta" + summary).textContent = details.length ? details.join(" • ") : emptyText;
    }

    function syncPlatformDefaults(type = 1) {
        if (type === 1 && byId("pass_platform")) {
            byId("pass_platform").value = value("platform") || "1";
        }
        invalidatePreparedAudio();
        schedulePrepareAnnouncement();
    }

    function selectAnnouncement(index, button) {
        selectedAnnouncement = index;
        document.querySelectorAll(".announce-option").forEach(btn => btn.classList.remove("active"));
        button.classList.add("active");
        refreshPlaybackControls();
        byId("selectedType").innerHTML = `<b>${escapeHtml(button.dataset.title || "ประเภทประกาศ")}</b><br>พร้อมสร้างเสียงตามข้อมูลที่เลือก`;

        ["delayFields", "passFields", "customFields"].forEach(id => byId(id).classList.remove("show"));
        byId("trainTypeWrap").classList.remove("hidden");
        if (index === 5) byId("delayFields").classList.add("show");
        if (index === 3 || index === 9) {
            byId("passFields").classList.add("show");
            if (index === 3) byId("trainTypeWrap").classList.add("hidden");
        }
        if (index === 8) {
            byId("customFields").classList.add("show");
            updateCustomLanguageFields();
        }
        byId("previewBox").innerHTML = "<b>กำลังเตรียมเสียงล่วงหน้า</b><br><br>เมื่อเสียงพร้อม ปุ่มประกาศจะทำงานได้แทบจะทันที";
        invalidatePreparedAudio();
        schedulePrepareAnnouncement(80);
    }

    function escapeHtml(text) {
        const div = document.createElement("div");
        div.textContent = text == null ? "" : String(text);
        return div.innerHTML;
    }

    function renderServerPreview(html) {
        const safe = escapeHtml(html).replace(/&lt;br\s*\/?&gt;/gi, "<br>");
        byId("previewBox").innerHTML = safe;
    }

    function collectPayload(tabIndex) {
        return {
            tab_index: tabIndex,
            announce_mode: value("announce_mode") || "thai_only",
            thai_voice: value("thai_voice") || "th-TH-PremwadeeNeural",
            num: value("num"), origin: value("origin"), dest: value("dest"), time: value("time"),
            platform: value("platform") || "1", current: value("current") || "คลองบางพระ",
            next: value("next_station"), delay: value("delay_time"),
            custom_text: value("custom_text"), custom_text_en: value("custom_text_en"),
            train_type: value("train_type") || "สินค้า",
            pass_platform: value("pass_platform") || value("platform") || "1",
            num_2: value("num_2"), origin_2: value("origin_2"), dest_2: value("dest_2"),
            time_2: value("time_2"), platform_2: value("platform_2"), next_2: value("next_station_2"),
            num_3: value("num_3"), origin_3: value("origin_3"), dest_3: value("dest_3"),
            time_3: value("time_3"), platform_3: value("platform_3"), next_3: value("next_station_3")
        };
    }

    function payloadKey(payload) {
        return JSON.stringify(payload);
    }

    function invalidatePreparedAudio() {
        preparedAudioKey = "";
        preparedAudioPromise = null;
        preparedAudioData = null;
        if (prepareTimer) {
            clearTimeout(prepareTimer);
            prepareTimer = null;
        }
    }

    function requestAnnouncementData(tabIndex, background = false) {
        const payload = collectPayload(tabIndex);
        const key = payloadKey(payload);

        if (preparedAudioKey === key && preparedAudioData) {
            return Promise.resolve(preparedAudioData);
        }
        if (preparedAudioKey === key && preparedAudioPromise) {
            return preparedAudioPromise;
        }

        preparedAudioKey = key;
        preparedAudioData = null;

        const requestPromise = fetch("/announce", {
            method: "POST",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify(payload)
        }).then(async response => {
            const data = await response.json();
            if (!response.ok || data.status !== "success") {
                throw new Error(data.message || "สร้างเสียงไม่สำเร็จ");
            }
            return data;
        }).then(data => {
            if (preparedAudioKey === key) {
                preparedAudioData = data;
                preparedAudioPromise = null;
                if (background && !isBusy) {
                    setStatus("เสียงพร้อมประกาศ", "ok");
                    byId("previewBox").innerHTML = "<b>เสียงพร้อมแล้ว</b><br><br>กดปุ่มเริ่มประกาศ เสียงเตือนจะดังทันที";
                }
            }
            return data;
        }).catch(error => {
            if (preparedAudioKey === key) {
                preparedAudioPromise = null;
                preparedAudioData = null;
            }
            throw error;
        });

        preparedAudioPromise = requestPromise;
        return requestPromise;
    }

    function schedulePrepareAnnouncement(delay = 320) {
        if (prepareTimer) clearTimeout(prepareTimer);
        if (selectedAnnouncement === null) return;

        prepareTimer = setTimeout(() => {
            prepareTimer = null;
            if (validateSelection()) return;
            if (!isBusy) setStatus("กำลังเตรียมเสียงล่วงหน้า...", "work");
            requestAnnouncementData(selectedAnnouncement, true).catch(error => {
                console.warn("Background audio preparation failed:", error);
                if (!isBusy) setStatus("พร้อมใช้งาน");
            });
        }, delay);
    }

    function validateSelection() {
        if (selectedAnnouncement === null) return "กรุณาเลือกประเภทประกาศ";
        if (![7, 8].includes(selectedAnnouncement) && !value("num") && ![3, 9].includes(selectedAnnouncement)) return "กรุณาเลือกขบวนรถ";
        if (selectedAnnouncement === 5 && !value("delay_time")) return "กรุณาระบุเวลาที่คาดว่าจะถึง";
        if (selectedAnnouncement === 8) {
            const mode = value("announce_mode");
            if (mode !== "english_only" && !value("custom_text")) return "กรุณาพิมพ์ข้อความภาษาไทย";
            if (mode !== "thai_only" && !value("custom_text_en")) return "กรุณาพิมพ์ข้อความภาษาอังกฤษ";
        }
        if (selectedAnnouncement === 10 && !value("num_2")) return "กรุณาเลือกอย่างน้อยขบวนที่ 1 และขบวนที่ 2";
        return "";
    }

    function setStatus(text, type = "normal") {
        const el = byId("statusText");
        el.textContent = text;
        if (type === "ok") el.style.background = "rgba(23,119,68,.92)";
        else if (type === "error") el.style.background = "rgba(180,35,24,.92)";
        else if (type === "work") el.style.background = "rgba(199,161,42,.96)";
        else el.style.background = "rgba(255,255,255,.12)";
    }

    function refreshPlaybackControls() {
        const playButton = byId("playButton");
        const pauseButton = byId("pauseButton");
        const stopButton = byId("stopButton");
        if (!playButton || !pauseButton || !stopButton) return;

        playButton.textContent = playbackState === "paused" ? "▶ เล่นต่อ" : "▶ เริ่มประกาศ";
        playButton.disabled = playbackState === "loading" || playbackState === "playing" || (playbackState === "idle" && selectedAnnouncement === null);
        pauseButton.disabled = playbackState !== "playing";
        stopButton.disabled = !["loading", "playing", "paused"].includes(playbackState);
    }

    function setPlaybackState(state) {
        playbackState = state;
        refreshPlaybackControls();
    }

    function setLoading(active) {
        byId("loadingBar").classList.toggle("active", active);
        document.querySelectorAll(".announce-option, .lang-btn, .voice-choice-btn, .voice-test-btn").forEach(btn => btn.disabled = active);
        refreshPlaybackControls();
    }

    function getMainPlayer() {
        if (!mainPlayer) {
            mainPlayer = new Audio();
            mainPlayer.preload = "auto";
            mainPlayer.setAttribute("playsinline", "");
            mainPlayer.setAttribute("webkit-playsinline", "");
        }
        return mainPlayer;
    }

    function makePlaybackStoppedError() {
        const error = new Error("หยุดการเล่นเสียงแล้ว");
        error.name = "PlaybackStoppedError";
        return error;
    }

    function cancelActivePlayback(resetPosition = true) {
        const cancel = activePlaybackCancel;
        activePlaybackCancel = null;
        if (cancel) {
            try { cancel(); } catch (e) {}
        }
        const player = getMainPlayer();
        try {
            player.pause();
            if (resetPosition) player.currentTime = 0;
        } catch (e) {}
    }

    function beginPlaybackRun() {
        playbackRunId += 1;
        cancelActivePlayback(true);
        setPlaybackState("loading");
        return playbackRunId;
    }

    async function testStationVoice() {
        if (isBusy) return;
        const runId = beginPlaybackRun();
        isBusy = true;
        await unlockMobileAudio();
        if (runId !== playbackRunId) return;
        setLoading(true);
        setStatus("กำลังสร้างเสียงทดสอบ...", "work");
        const testMode = value("announce_mode") || "thai_only";
        byId("previewBox").innerHTML = "<b>กำลังทดสอบเสียง</b><br><br>ระบบกำลังสร้างเสียงตามภาษาและเพศที่เลือก";

        try {
            const response = await fetch("/test-station-voice", {
                method: "POST",
                headers: { "Content-Type": "application/json" },
                body: JSON.stringify({
                    thai_voice: value("thai_voice") || "th-TH-PremwadeeNeural",
                    announce_mode: testMode
                })
            });
            if (runId !== playbackRunId) throw makePlaybackStoppedError();
            const data = await response.json();
            if (!response.ok || data.status !== "success") {
                throw new Error(data.message || "สร้างเสียงทดสอบไม่สำเร็จ");
            }
            const testUrls = (data.audio_urls && data.audio_urls.length) ? data.audio_urls : [data.audio_url].filter(Boolean);
            const testLabels = data.audio_labels || [];
            if (!testUrls.length) throw new Error("ไม่พบไฟล์เสียงทดสอบ");
            for (let i = 0; i < testUrls.length; i++) {
                if (runId !== playbackRunId) throw makePlaybackStoppedError();
                setStatus(`กำลังทดสอบ ${testLabels[i] || "เสียง"}`, "ok");
                await playUrl(testUrls[i], { errorText: "ไม่สามารถเล่นเสียงทดสอบได้", runId });
            }
            if (runId !== playbackRunId) throw makePlaybackStoppedError();
            renderServerPreview(data.text_preview || "-");
            setStatus("ทดสอบเสียงเสร็จแล้ว", "ok");
        } catch (err) {
            if (err?.name !== "PlaybackStoppedError" && runId === playbackRunId) {
                console.error(err);
                setStatus("เกิดข้อผิดพลาด", "error");
                byId("previewBox").innerHTML = `<b>เกิดข้อผิดพลาด</b><br><br>${escapeHtml(err.message || String(err))}`;
            }
        } finally {
            if (runId === playbackRunId) {
                setLoading(false);
                isBusy = false;
                setPlaybackState("idle");
            }
        }
    }

    async function playOrResumeAudio() {
        const player = getMainPlayer();
        if (playbackState === "paused") {
            try {
                await player.play();
                setPlaybackState("playing");
                setStatus("เล่นเสียงต่อแล้ว", "ok");
                logHistoryEvent("resume");
            } catch (err) {
                setStatus("เล่นเสียงต่อไม่สำเร็จ", "error");
            }
            return;
        }
        await playSelectedAnnouncement();
    }

    function pauseAudio() {
        if (playbackState !== "playing") return;
        const player = getMainPlayer();
        try {
            player.pause();
            setPlaybackState("paused");
            setStatus("พักเสียงชั่วคราว", "work");
            logHistoryEvent("pause");
        } catch (e) {}
    }

    function stopAudio() {
        if (["loading", "playing", "paused"].includes(playbackState)) {
            logHistoryEvent("stop");
        }
        playbackRunId += 1;
        cancelActivePlayback(true);
        isBusy = false;
        setLoading(false);
        setPlaybackState("idle");
        setStatus("หยุดเสียงแล้ว");
        activeHistoryId = null;
        activeHistoryPromise = null;
    }

    function getAudioContext() {
        try {
            const AudioContext = window.AudioContext || window.webkitAudioContext;
            if (!sharedAudioContext) sharedAudioContext = new AudioContext();
            if (sharedAudioContext.state === "suspended") sharedAudioContext.resume().catch(() => {});
            return sharedAudioContext;
        } catch (e) { return null; }
    }

    async function unlockMobileAudio() {
        getAudioContext();
        if (mobileAudioUnlocked) return;
        const player = getMainPlayer();
        try {
            player.muted = true; player.volume = 0;
            player.src = "/audio/chime.mp3";
            player.currentTime = 0;
            await player.play();
            player.pause();
            try { player.currentTime = 0; } catch (e) {}
            mobileAudioUnlocked = true;
        } catch (e) { console.warn("Mobile audio unlock failed:", e); }
        finally { player.muted = false; player.volume = 1; }
    }

    function playWarningTone() {
        return new Promise(resolve => {
            try {
                const ctx = getAudioContext();
                if (!ctx) return resolve();
                const startAt = ctx.currentTime + .02;
                const gain = ctx.createGain();
                gain.gain.setValueAtTime(.0001, startAt);
                gain.gain.exponentialRampToValueAtTime(.12, startAt + .02);
                gain.gain.exponentialRampToValueAtTime(.0001, startAt + .42);
                gain.connect(ctx.destination);
                const osc1 = ctx.createOscillator(); osc1.type = "sine"; osc1.frequency.setValueAtTime(880, startAt); osc1.connect(gain); osc1.start(startAt); osc1.stop(startAt + .18);
                const osc2 = ctx.createOscillator(); osc2.type = "sine"; osc2.frequency.setValueAtTime(660, startAt + .19); osc2.connect(gain); osc2.start(startAt + .19); osc2.stop(startAt + .42);
                setTimeout(resolve, 455);
            } catch (e) { resolve(); }
        });
    }

    function playUrl(url, options = {}) {
        return new Promise((resolve, reject) => {
            const player = getMainPlayer();
            const maxWaitMs = options.maxWaitMs || null;
            const errorText = options.errorText || "เล่นไฟล์เสียงไม่สำเร็จ";
            const runId = options.runId ?? playbackRunId;
            let finished = false, started = false, safetyTimer = null, startTimer = null;

            function cleanup() {
                player.removeEventListener("ended", onEnded);
                player.removeEventListener("error", onError);
                player.removeEventListener("canplay", startPlay);
                if (safetyTimer) clearTimeout(safetyTimer);
                if (startTimer) clearTimeout(startTimer);
                if (activePlaybackCancel === cancelThisPlayback) activePlaybackCancel = null;
            }
            function finish() {
                if (finished) return;
                finished = true;
                cleanup();
                if (maxWaitMs) {
                    try { player.pause(); player.currentTime = 0; } catch (e) {}
                }
                resolve();
            }
            function fail(e) {
                if (finished) return;
                finished = true;
                cleanup();
                reject(e instanceof Error ? e : new Error(errorText));
            }
            function cancelThisPlayback() {
                if (finished) return;
                try { player.pause(); player.currentTime = 0; } catch (e) {}
                fail(makePlaybackStoppedError());
            }
            function onEnded() { finish(); }
            function onError() { fail(new Error(errorText)); }
            async function startPlay() {
                if (started || finished) return;
                if (runId !== playbackRunId) return fail(makePlaybackStoppedError());
                started = true;
                try {
                    player.muted = false;
                    player.volume = 1;
                    player.playbackRate = 1.0;
                    await player.play();
                    setPlaybackState("playing");
                } catch (e) { fail(e); }
            }
            try {
                if (runId !== playbackRunId) return fail(makePlaybackStoppedError());
                cancelActivePlayback(true);
                activePlaybackCancel = cancelThisPlayback;
                player.src = url;
                player.preload = "auto";
                player.muted = false;
                player.volume = 1;
                player.currentTime = 0;
                player.addEventListener("ended", onEnded, { once: true });
                player.addEventListener("error", onError, { once: true });
                player.addEventListener("canplay", startPlay, { once: true });
                player.load();
                startTimer = setTimeout(startPlay, 450);
                if (maxWaitMs) safetyTimer = setTimeout(finish, maxWaitMs);
            } catch (e) { fail(e); }
        });
    }

    async function playOriginalChime(runId = playbackRunId) {
        try {
            await playUrl("/audio/chime.mp3", { maxWaitMs: 5200, errorText: "เล่นเสียงเตือนไม่สำเร็จ", runId });
        } catch (e) {
            if (e?.name === "PlaybackStoppedError") throw e;
            if (runId !== playbackRunId) throw makePlaybackStoppedError();
            await playWarningTone();
        }
    }

    async function playSelectedAnnouncement() {
        const error = validateSelection();
        if (error) {
            setStatus("ข้อมูลไม่ครบ", "error");
            byId("previewBox").innerHTML = `<b>กรุณาตรวจสอบข้อมูล</b><br><br>${escapeHtml(error)}`;
            return;
        }
        await playAnnouncement(selectedAnnouncement);
    }

    async function playAnnouncement(tabIndex) {
        if (isBusy) return;
        const runId = beginPlaybackRun();
        isBusy = true;
        activeHistoryId = null;
        activeHistoryPromise = startHistoryRecord(tabIndex);
        await unlockMobileAudio();
        if (runId !== playbackRunId) return;
        setLoading(true);

        // เริ่มสร้าง/ดึงไฟล์เสียงและเล่นเสียงเตือนพร้อมกัน
        setStatus("เสียงเตือน...", "work");
        byId("previewBox").innerHTML = "<b>กำลังเริ่มประกาศ</b><br><br>เสียงเตือนกำลังเล่น และระบบกำลังตรวจสอบไฟล์เสียงประกาศ";

        try {
            const dataPromise = requestAnnouncementData(tabIndex, false);
            const chimePromise = playOriginalChime(runId);
            const [data, , historyId] = await Promise.all([dataPromise, chimePromise, activeHistoryPromise]);
            if (runId !== playbackRunId) throw makePlaybackStoppedError();
            activeHistoryId = historyId;
            logHistoryEvent("generated", {
                message: data.text_preview || "",
                generation_ms: data.generation_ms || 0
            });

            renderServerPreview(data.text_preview || "-");
            const audioUrls = (data.audio_urls && data.audio_urls.length) ? data.audio_urls : [data.audio_url].filter(Boolean);
            const audioLabels = data.audio_labels || [];
            if (!audioUrls.length) throw new Error("ไม่พบไฟล์เสียงสำหรับประกาศ");

            audioUrls.forEach(url => {
                try { fetch(url, { cache: "force-cache" }).catch(() => {}); } catch (e) {}
            });

            for (let i = 0; i < audioUrls.length; i++) {
                if (runId !== playbackRunId) throw makePlaybackStoppedError();
                setStatus(`กำลังประกาศ ${audioLabels[i] || ""}`.trim(), "ok");
                await playUrl(audioUrls[i], {
                    errorText: "มือถือบล็อกเสียงประกาศ กรุณากดปุ่มอีกครั้ง",
                    runId
                });
            }
            if (runId !== playbackRunId) throw makePlaybackStoppedError();
            setStatus("ประกาศเสร็จแล้ว", "ok");
            await logHistoryEvent("success");
        } catch (err) {
            if (err?.name !== "PlaybackStoppedError" && runId === playbackRunId) {
                console.error(err);
                logHistoryEvent("failed", { reason: err.message || String(err) });
                setStatus("เกิดข้อผิดพลาด", "error");
                let message = err.message || String(err);
                if (message.includes("not allowed") || message.includes("permission") || message.includes("บล็อก")) {
                    message = "มือถือบล็อกการเล่นเสียงชั่วคราว กรุณากดปุ่มประกาศอีกครั้ง หรือเปิดหน้านี้ผ่าน Chrome/Safari โดยตรง";
                }
                byId("previewBox").innerHTML = `<b>เกิดข้อผิดพลาด</b><br><br>${escapeHtml(message)}`;
            }
        } finally {
            if (runId === playbackRunId) {
                setLoading(false);
                isBusy = false;
                setPlaybackState("idle");
                activeHistoryId = null;
                activeHistoryPromise = null;
            }
        }
    }

    function clearData() {
        stopAudio();
        invalidatePreparedAudio();
        ["train_select", "num", "time", "origin", "dest", "next_station", "delay_time", "custom_text", "custom_text_en",
         "train_select_2", "num_2", "time_2", "origin_2", "dest_2", "next_station_2",
         "train_select_3", "num_3", "time_3", "origin_3", "dest_3", "next_station_3"].forEach(id => { if (byId(id)) byId(id).value = ""; });
        byId("platform").value = "1"; byId("pass_platform").value = "1"; byId("platform_2").value = "2"; byId("platform_3").value = "3";
        byId("current").value = "คลองบางพระ"; byId("train_type").value = "สินค้า";
        setLanguage("thai_only", document.querySelector('[data-mode="thai_only"]'));
        selectedAnnouncement = null;
        document.querySelectorAll(".announce-option").forEach(btn => btn.classList.remove("active"));
        ["delayFields", "passFields", "customFields"].forEach(id => byId(id).classList.remove("show"));
        refreshPlaybackControls();
        byId("selectedType").innerHTML = "<b>ยังไม่ได้เลือกประเภทประกาศ</b><br>เลือกปุ่มในขั้นตอนที่ 3 ก่อน";
        byId("previewBox").innerHTML = "<b>ตัวอย่างข้อความประกาศ</b><br><br>เมื่อกดเริ่มประกาศ ระบบจะสร้างข้อความและไฟล์เสียงตามภาษาที่เลือก";
        [1, 2, 3].forEach(refreshSummary); setStatus("พร้อมใช้งาน");
    }

    // เมื่อแก้ข้อมูลหลังเลือกประเภทประกาศ ให้เตรียมเสียงชุดใหม่อัตโนมัติ
    document.querySelectorAll("input, select, textarea").forEach(element => {
        element.addEventListener("input", () => {
            invalidatePreparedAudio();
            schedulePrepareAnnouncement();
        });
        element.addEventListener("change", () => {
            invalidatePreparedAudio();
            schedulePrepareAnnouncement(180);
        });
    });

    updateCustomLanguageFields();
    [1, 2, 3].forEach(refreshSummary);
    refreshPlaybackControls();
</script>
</body>
</html>
"""



def group_buttons(buttons):
    grouped = {}
    for item in buttons:
        grouped.setdefault(item["group"], []).append(item)
    return grouped


def cleanup_old_audio(max_age_seconds=AUDIO_CACHE_MAX_AGE):
    """ลบไฟล์เสียงเก่าแบบเป็นช่วง ๆ เพื่อลดภาระทุกครั้งที่กดประกาศ"""
    global _LAST_AUDIO_CLEANUP
    now = time.time()
    if now - _LAST_AUDIO_CLEANUP < 600:
        return

    with _AUDIO_CLEANUP_LOCK:
        now = time.time()
        if now - _LAST_AUDIO_CLEANUP < 600:
            return
        _LAST_AUDIO_CLEANUP = now

        for file in AUDIO_DIR.glob("announce_*.mp3"):
            try:
                if now - file.stat().st_mtime > max_age_seconds:
                    file.unlink(missing_ok=True)
            except OSError:
                pass


def _cache_lock(cache_key):
    with _AUDIO_CACHE_LOCKS_GUARD:
        return _AUDIO_CACHE_LOCKS.setdefault(cache_key, threading.Lock())


def generate_cached_audio(segment):
    """สร้างไฟล์ TTS ครั้งเดียว แล้วใช้ซ้ำจากแคชตามข้อความและเสียงที่ตรงกัน"""
    prepare_func = segment.get("prepare") or (lambda value: value)
    tts_text = prepare_func(segment.get("text", ""))
    cache_payload = json.dumps(
        {
            "voice": segment["voice"],
            "rate": segment["rate"],
            "volume": segment["volume"],
            "pitch": segment["pitch"],
            "text": tts_text,
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    cache_key = hashlib.sha256(cache_payload.encode("utf-8")).hexdigest()[:28]
    filename = f"announce_cache_{segment['code']}_{cache_key}.mp3"
    output_path = AUDIO_DIR / filename

    if output_path.exists() and output_path.stat().st_size > 0:
        try:
            os.utime(output_path, None)
        except OSError:
            pass
        return filename

    lock = _cache_lock(cache_key)
    with lock:
        if output_path.exists() and output_path.stat().st_size > 0:
            try:
                os.utime(output_path, None)
            except OSError:
                pass
            return filename

        temp_path = AUDIO_DIR / f"announce_tmp_{uuid.uuid4().hex}.mp3"
        try:
            subprocess.run(
                [
                    "edge-tts",
                    f"--voice={segment['voice']}",
                    f"--rate={segment['rate']}",
                    f"--volume={segment['volume']}",
                    f"--pitch={segment['pitch']}",
                    "--text", tts_text,
                    "--write-media", str(temp_path),
                ],
                check=True,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
            if not temp_path.exists() or temp_path.stat().st_size == 0:
                raise RuntimeError(f"ระบบสร้างไฟล์เสียงช่วง {segment['code']} ไม่สำเร็จ หรือไฟล์เสียงว่าง")
            os.replace(temp_path, output_path)
        finally:
            temp_path.unlink(missing_ok=True)

    return filename


def generate_audio_segments(segments):
    """สร้างไทยและอังกฤษพร้อมกัน ลดเวลาของโหมดสองภาษาเกือบครึ่งหนึ่ง"""
    if not segments:
        return []
    if len(segments) == 1:
        return [generate_cached_audio(segments[0])]

    with ThreadPoolExecutor(max_workers=min(2, len(segments))) as executor:
        return list(executor.map(generate_cached_audio, segments))


def spaced_train_number(number_text):
    digits = "".join(ch for ch in str(number_text) if ch.isdigit())
    return " ".join(digits) if digits else str(number_text or "")


def tidy_time(text):
    return (text or "").replace(" 0 นาที", "").strip()


def station(name):
    name = (name or "").strip()
    if not name:
        return ""
    if name.startswith("สถานี") or name.startswith("ป้ายหยุดรถ") or name.startswith("ที่หยุดรถ"):
        return name
    return f"สถานี{name}"


def clean_space(text):
    return " ".join((text or "").split())


def english_voice_for(thai_voice):
    """เลือกเสียงอังกฤษให้เพศตรงกับเสียงไทยที่ผู้ใช้เลือก"""
    if thai_voice == "th-TH-NiwatNeural":
        return ENGLISH_VOICE_OPTIONS["male"]
    return ENGLISH_VOICE_OPTIONS["female"]


# แก้เฉพาะคำที่ระบบ TTS มักอ่านคลาดเคลื่อนจริง ๆ
# ข้อความที่แสดงบนหน้าเว็บจะยังคงแพตเทิร์นเดิมทุกคำ
PRONUNCIATION_FIXES = {
    "กบินทร์บุรี": "กะบินบุรี",
    "จุกเสม็ด": "จุก สะ เหม็ด",
    # ขยายคำย่อเฉพาะตอนอ่านเสียง เพื่อไม่ให้ TTS อ่านเครื่องหมาย ฯ แปลก ๆ
    "การรถไฟฯ": "การรถไฟแห่งประเทศไทย",
}


def _normalize_punctuation_spacing(text):
    """จัดช่องว่างรอบเครื่องหมาย โดยไม่รบกวนเวลา 8:25 หรือคำย่อ P.M."""
    value = text or ""
    value = re.sub(r"\s+([,;!?])", r"\1", value)
    value = re.sub(r"([,;!?])(?=[^\s])", r"\1 ", value)
    value = re.sub(r",\s*,+", ",", value)
    value = re.sub(r"\.\s*\.+", ".", value)
    return clean_space(value)


def prepare_tts_text(text):
    """
    เตรียมข้อความภาษาไทยสำหรับเสียงเท่านั้น

    หลักการคือคงรูปประโยคประกาศเดิม แต่เพิ่มเครื่องหมายหยุดในตำแหน่ง
    ที่ผู้ประกาศจริงมักเว้นหายใจ เพื่อไม่ให้เสียงอ่านรวดเดียวติดกัน
    """
    tts_text = clean_space(text or "")

    for official_word, spoken_word in PRONUNCIATION_FIXES.items():
        tts_text = tts_text.replace(official_word, spoken_word)

    # เปิดประกาศให้มีจังหวะก่อนเข้าสู่รายละเอียด
    opening_rules = (
        (r"^ท่านผู้โดยสารโปรดทราบ\s+", "ท่านผู้โดยสารโปรดทราบ. "),
        (r"^ผู้โดยสารโปรดทราบ\s+", "ผู้โดยสารโปรดทราบ. "),
        (r"^โปรดทราบ\s+", "โปรดทราบ. "),
        (r"\bอีกสักครู่\s+", "อีกสักครู่, "),
        (r"\bวันนี้ขบวนรถ\s+", "วันนี้, ขบวนรถ "),
    )
    for pattern, replacement in opening_rules:
        tts_text = re.sub(pattern, replacement, tts_text, count=1)

    # คำว่า “ที่นี่” และชื่อสถานีต้องเป็นวลีเดียวกัน เพื่อไม่ให้เสียงหยุดกลางวลี
    # กล่าวชื่อสถานีซ้ำด้วยจังหวะสั้นแบบผู้ประกาศจริง ไม่ใช้จุดเต็มประโยค
    tts_text = re.sub(
        r"(ที่นี่สถานี.+?)\s+(?=ที่นี่สถานี)",
        r"\1, ",
        tts_text,
        count=1,
    )

    # หลังกล่าวชื่อสถานีครั้งที่สอง ให้พักเพียงเล็กน้อยก่อนแจ้งผู้โดยสาร
    tts_text = re.sub(
        r"(ที่นี่สถานี.+?)(?=\s+(?:ผู้โดยสารก่อน|ก่อนผู้โดยสาร))",
        r"\1, ",
        tts_text,
        count=1,
    )

    # จังหวะข้อมูลขบวน: หมายเลข → ต้นทาง → ปลายทาง → เวลา
    tts_text = re.sub(
        r"(ขบวนที่\s+(?:\d\s*)+)(?=\s)",
        lambda m: m.group(1).rstrip() + ", ",
        tts_text,
    )
    tts_text = tts_text.replace("ขบวนรถ ขบวนที่", "ขบวนรถ, ขบวนที่")
    tts_text = tts_text.replace("รับส่งผู้โดยสารต้นทาง", "รับส่งผู้โดยสาร, ต้นทาง")
    tts_text = re.sub(r"\s+ปลายทาง\s+", ", ปลายทาง ", tts_text)
    tts_text = re.sub(r"\s+เที่ยวกำหนดเวลา\s+", ", เที่ยวกำหนดเวลา ", tts_text)
    tts_text = tts_text.replace("ต้นทาง สถานี", "ต้นทางสถานี")
    tts_text = tts_text.replace("ปลายทาง สถานี", "ปลายทางสถานี")

    # หลังบอกเวลา ให้จบช่วงข้อมูลขบวนก่อนเข้าสู่คำแนะนำถัดไป
    tts_text = re.sub(
        r"(เที่ยวกำหนดเวลา\s+\d+\s+นาฬิกา(?:\s+\d+\s+นาที)?)(?=\s)",
        r"\1. ",
        tts_text,
    )

    # หัวข้อความปลอดภัยและข้อความขอความร่วมมือ
    tts_text = re.sub(r"\s+เพื่อความปลอดภัย\s+", ". เพื่อความปลอดภัย, ", tts_text)
    tts_text = re.sub(r"\s+เพื่อความปลอดภัยและสุขอนามัยที่ดี\s+", ". เพื่อความปลอดภัยและสุขอนามัยที่ดี, ", tts_text)
    tts_text = tts_text.replace(" และไม่เดินข้ามไปมา", ", และไม่เดินข้ามไปมา")
    tts_text = tts_text.replace(" ระหว่างชานชาลาที่", ", ระหว่างชานชาลาที่")

    # จังหวะเมื่อกล่าวถึงชานชาลา: ใช้จังหวะสั้นถ้ายังเป็นประโยคเดียวกัน
    tts_text = re.sub(
        r"(ชานชาลาที่\s+\d+)(?=\s+(?:เป็นขบวนรถ|เมื่อออกจาก))",
        r"\1, ",
        tts_text,
    )
    tts_text = re.sub(
        r"(ชานชาลาที่\s+\d+)(?=\s+(?:เพื่อความปลอดภัย|ผู้โดยสาร|ขอบคุณ|โปรด))",
        r"\1. ",
        tts_text,
    )

    # แบ่งช่วงคำแนะนำผู้โดยสารและข้อมูลขบวนให้ไม่อ่านรวดเดียว
    tts_text = tts_text.replace("ผู้โดยสารก่อนลงจากขบวนรถ โปรดตรวจสอบ", "ผู้โดยสารก่อนลงจากขบวนรถ, โปรดตรวจสอบ")
    tts_text = tts_text.replace("ก่อนผู้โดยสารจะลงจากขบวนรถ โปรดตรวจสอบ", "ก่อนผู้โดยสารจะลงจากขบวนรถ, โปรดตรวจสอบ")
    tts_text = tts_text.replace("สิ่งของและสัมภาระของท่าน นำลง", "สิ่งของและสัมภาระของท่าน, นำลง")
    tts_text = tts_text.replace("นำลงจากขบวนรถให้ครบถ้วน ขบวนรถ", "นำลงจากขบวนรถให้ครบถ้วน. ขบวนรถ")
    tts_text = tts_text.replace("นำลงให้ถูกต้องครบถ้วน ขบวนรถ", "นำลงให้ถูกต้องครบถ้วน. ขบวนรถ")
    tts_text = tts_text.replace("ผู้โดยสารที่ลงจากขบวนรถ โปรดระมัดระวัง", "ผู้โดยสารที่ลงจากขบวนรถ, โปรดระมัดระวัง")

    # ประกาศหลายขบวน: เว้นจังหวะระหว่างรายละเอียดแต่ละขบวน
    tts_text = re.sub(r"\s+(?=และขบวนรถที่จอดในชานชาลาที่)", ", ", tts_text)
    tts_text = re.sub(r"\s+(?=ขบวนรถในชานชาลาที่)", ". ", tts_text, count=1)
    tts_text = re.sub(r"\s+(?=และขบวนรถในชานชาลาที่)", ", ", tts_text)

    # แยกข้อความสถานีต่อไปและคำขอโทษให้ฟังเป็นประโยคชัดเจน
    tts_text = tts_text.replace("ที่ ป้ายหยุดรถ", "ที่ป้ายหยุดรถ")
    tts_text = tts_text.replace("ที่ สถานี", "ที่สถานี")
    tts_text = tts_text.replace("สถานีคลองบางพระ แล้ว", "สถานีคลองบางพระแล้ว")
    tts_text = tts_text.replace(" และ สถานี", " และสถานี")
    tts_text = tts_text.replace(" และ ป้ายหยุดรถ", " และป้ายหยุดรถ")
    tts_text = tts_text.replace(" เป็นสถานีต่อไปตามลำดับ", ", เป็นสถานีต่อไปตามลำดับ")
    tts_text = tts_text.replace(" ล่าช้ากว่ากำหนดเวลาเดิม คาดว่าจะถึง", " ล่าช้ากว่ากำหนดเวลาเดิม. คาดว่าจะถึง")
    tts_text = tts_text.replace(" ในนามของการรถไฟแห่งประเทศไทย", ". ในนามของการรถไฟแห่งประเทศไทย")
    tts_text = tts_text.replace(" ต้องขออภัย", ", ต้องขออภัย")

    # ประกาศห้ามสูบบุหรี่มีหลายใจความ จึงแบ่งเป็นช่วงสั้น ๆ
    tts_text = tts_text.replace("ขอแจ้งให้ทราบว่า บริเวณสถานี", "ขอแจ้งให้ทราบว่า, บริเวณสถานี")
    tts_text = tts_text.replace("ภายในเขตพื้นที่สถานีทุกแห่ง เป็นเขต", "ภายในเขตพื้นที่สถานีทุกแห่ง, เป็นเขต")
    tts_text = tts_text.replace("เครื่องดื่มแอลกอฮอล์ ห้ามสูบบุหรี่", "เครื่องดื่มแอลกอฮอล์. ห้ามสูบบุหรี่")
    tts_text = tts_text.replace("โดยเด็ดขาด ผู้ฝ่าฝืน", "โดยเด็ดขาด. ผู้ฝ่าฝืน")

    # คำลงท้ายควรมีจังหวะก่อนกล่าวขอบคุณ แต่ไม่เพิ่มถ้ามีเครื่องหมายอยู่แล้ว
    tts_text = re.sub(r"(?<![.!?])\s+(ขอขอบคุณในความร่วมมือ(?:ครับ|ค่ะ))$", r". \1", tts_text)
    tts_text = re.sub(r"(?<![.!?])\s+(ขอบคุณ(?:ครับ|ค่ะ))$", r". \1", tts_text)

    return _normalize_punctuation_spacing(tts_text)


def prepare_english_tts_text(text):
    """จัดช่องว่างภาษาอังกฤษโดยรักษาเวลา 8:25 และคำย่อ A.M./P.M."""
    tts_text = clean_space(text or "")
    tts_text = re.sub(r"\s+([,.;:!?])", r"\1", tts_text)
    # ป้องกันเสียงอ่าน Attention please ติดกับประโยคถัดไปเร็วเกินไป
    tts_text = re.sub(r"^Attention please[,.]?\s*", "Attention please. ", tts_text, count=1, flags=re.IGNORECASE)
    tts_text = tts_text.replace("A.M..", "A.M.").replace("P.M..", "P.M.")
    return clean_space(tts_text)


def apply_voice_politeness(text, thai_voice):
    """เปลี่ยนเฉพาะคำลงท้ายตามเพศของเสียง โดยไม่เปลี่ยนโครงสร้างประโยค"""
    result = text or ""
    if thai_voice == "th-TH-PremwadeeNeural":
        return result.replace("ครับ", "ค่ะ")
    return result.replace("ค่ะ", "ครับ")



EN_STATION_NAMES = {
    "คลองบางพระ": "Khlong Bang Phra",
    "กรุงเทพ": "Bangkok",
    "หัวลำโพง": "Hua Lamphong",
    "ชุมทางฉะเชิงเทรา": "Chachoengsao Junction",
    "ฉะเชิงเทรา": "Chachoengsao",
    "ปราจีนบุรี": "Prachin Buri",
    "กบินทร์บุรี": "Kabin Buri",
    "จุกเสม็ด": "Chuk Samet",
    "ด่านพรมแดนบ้านคลองลึก": "Ban Klong Luk Border",
    "คลองแขวงกลั่น": "Khlong Khwaeng Klan",
    "คลองเปรง": "Khlong Preng",
    "บางเตย": "Bang Toei",
    "สถานีคลองเปรง": "Khlong Preng Station",
    "สถานีชุมทางฉะเชิงเทรา": "Chachoengsao Junction Station",
    "ป้ายหยุดรถคลองแขวงกลั่น": "Khlong Khwaeng Klan Halt",
    "ป้ายหยุดรถบางเตย": "Bang Toei Halt",
}

EN_TRAIN_TYPES = {
    "สินค้า": "freight train",
    "ด่วนพิเศษ": "special express train",
    "พิเศษ": "special train",
    "รถจักรเปล่า": "light engine",
}

DIGIT_WORDS_EN = {
    "0": "zero", "1": "one", "2": "two", "3": "three", "4": "four",
    "5": "five", "6": "six", "7": "seven", "8": "eight", "9": "nine",
}


def train_number_en(number_text):
    digits = "".join(ch for ch in str(number_text) if ch.isdigit())
    return " ".join(DIGIT_WORDS_EN.get(ch, ch) for ch in digits) if digits else str(number_text or "")


def station_en(name):
    name = (name or "").strip()
    if not name:
        return ""
    cleaned = name
    for prefix in ("สถานี", "ป้ายหยุดรถ", "ที่หยุดรถ"):
        if cleaned.startswith(prefix):
            cleaned = cleaned.replace(prefix, "", 1).strip()
    return EN_STATION_NAMES.get(name) or EN_STATION_NAMES.get(cleaned) or cleaned


def next_stations_en(text):
    result = text or ""
    # แทนคำที่ยาวก่อน เพื่อไม่ให้ชนกับชื่อสถานีย่อย
    for th_name in sorted(EN_STATION_NAMES, key=len, reverse=True):
        result = result.replace(th_name, EN_STATION_NAMES[th_name])
    result = result.replace(" และ ", " and ").replace("และ", " and ")
    return clean_space(result)


def time_en(text):
    raw = text or ""
    nums = [int(part) for part in __import__("re").findall(r"\d+", raw)]
    if not nums:
        return raw
    hour = nums[0]
    minute = nums[1] if len(nums) > 1 else 0
    suffix = "A.M." if hour < 12 else "P.M."
    hour12 = hour % 12
    if hour12 == 0:
        hour12 = 12
    return f"{hour12}:{minute:02d} {suffix}"


def build_english_announcement(data):
    idx = int(data.get("tab_index", -1))

    t_num = train_number_en(data.get("num", ""))
    origin = station_en(data.get("origin", ""))
    dest = station_en(data.get("dest", ""))
    t_time = time_en(tidy_time(data.get("time", "")))
    platform = data.get("platform", "")
    pass_platform = data.get("pass_platform") or platform
    current = station_en(data.get("current", STATION_NAME) or STATION_NAME)
    next_st = next_stations_en(data.get("next", ""))
    delay = time_en(data.get("delay", ""))
    custom_text_en = data.get("custom_text_en", "")
    train_type = EN_TRAIN_TYPES.get(data.get("train_type", "สินค้า") or "สินค้า", "train")

    t_num_2 = train_number_en(data.get("num_2", ""))
    origin_2 = station_en(data.get("origin_2", ""))
    dest_2 = station_en(data.get("dest_2", ""))
    platform_2 = data.get("platform_2", "")
    next_st_2 = next_stations_en(data.get("next_2", ""))

    t_num_3 = train_number_en(data.get("num_3", ""))
    origin_3 = station_en(data.get("origin_3", ""))
    dest_3 = station_en(data.get("dest_3", ""))
    platform_3 = data.get("platform_3", "")
    next_st_3 = next_stations_en(data.get("next_3", ""))

    if idx == 0:
        text = f"Attention please. Passengers traveling on train number {t_num}, from {origin} to {dest}, scheduled at {t_time}, please purchase your ticket at the ticket office before boarding."
    elif idx == 1:
        text = f"Attention please. Passengers holding tickets for train number {t_num}, from {origin} to {dest}, scheduled at {t_time}, please wait with your belongings on platform {platform}."
    elif idx == 2:
        text = f"Attention please. Train number {t_num}, from {origin} to {dest}, scheduled at {t_time}, will shortly arrive at platform {platform}. For your safety, please stand behind the yellow line and do not cross the tracks."
    elif idx == 3:
        text = f"Attention please. A train will shortly pass through platform {pass_platform}. For your safety, please stand behind the yellow line and do not cross the tracks."
    elif idx == 4:
        text = f"Attention please. This is {current} Station. Before leaving the train, please check all your belongings. The train at platform {platform} is train number {t_num}, from {origin} to {dest}. After departing {current} Station, the next stops will be {next_st}."
    elif idx == 5:
        text = f"Attention please. Train number {t_num}, from {origin} to {dest}, scheduled at {t_time}, is delayed. The train is expected to arrive at {current} Station at approximately {delay}. The State Railway of Thailand apologizes for the inconvenience."
    elif idx == 6:
        text = f"Attention please. Train number {t_num} will shortly arrive at platform {platform}. Passengers leaving the train, please be careful."
    elif idx == 7:
        text = "Attention please. For safety and good hygiene, the State Railway of Thailand would like to inform all passengers that all station areas, trains, and railway station premises are smoke-free and alcohol-free areas. Smoking and drinking alcoholic beverages are strictly prohibited. Violators are subject to legal action."
    elif idx == 8:
        text = custom_text_en.strip() if custom_text_en.strip() else "Attention please. Please listen carefully to the station announcement."
    elif idx == 9:
        text = f"Attention please. A {train_type} will shortly pass through platform {pass_platform}. For your safety, please stand behind the yellow line and do not cross the tracks."
    elif idx == 10:
        trains = [
            (platform, t_num, origin, dest, next_st),
            (platform_2, t_num_2, origin_2, dest_2, next_st_2),
        ]
        if t_num_3:
            trains.append((platform_3, t_num_3, origin_3, dest_3, next_st_3))
        train_details = " ".join(
            f"The train at platform {p} is train number {n}, from {o} to {d}."
            for p, n, o, d, _ in trains
        )
        next_details = " ".join(
            f"After departing {current} Station, the train at platform {p} will next stop at {nxt}."
            for p, _, _, _, nxt in trains
        )
        text = (
            f"Attention please. This is {current} Station. Before leaving the train, please check all your belongings. "
            f"{train_details} {next_details}"
        )
    else:
        raise ValueError("ไม่พบประเภทประกาศที่เลือก")

    text = clean_space(text)
    if not text.lower().endswith("thank you."):
        text = f"{text} Thank you."
    return text

def build_announcement(data):
    idx = int(data.get("tab_index", -1))

    t_num = spaced_train_number(data.get("num", ""))
    origin = data.get("origin", "")
    dest = data.get("dest", "")
    t_time = tidy_time(data.get("time", ""))
    platform = data.get("platform", "")
    pass_platform = data.get("pass_platform") or platform
    current = data.get("current", STATION_NAME) or STATION_NAME
    next_st = data.get("next", "")
    delay = data.get("delay", "")
    custom_text = data.get("custom_text", "")
    train_type = data.get("train_type", "สินค้า") or "สินค้า"

    t_num_2 = spaced_train_number(data.get("num_2", ""))
    origin_2 = data.get("origin_2", "")
    dest_2 = data.get("dest_2", "")
    platform_2 = data.get("platform_2", "")
    next_st_2 = data.get("next_2", "")

    t_num_3 = spaced_train_number(data.get("num_3", ""))
    origin_3 = data.get("origin_3", "")
    dest_3 = data.get("dest_3", "")
    platform_3 = data.get("platform_3", "")
    next_st_3 = data.get("next_3", "")

    if idx == 0:
        text = f"โปรดทราบ ผู้โดยสารที่มีความประสงค์จะเดินทางกับขบวนรถ ขบวนที่ {t_num} รับส่งผู้โดยสารต้นทาง {station(origin)} ปลายทาง {station(dest)} เที่ยวกำหนดเวลา {t_time} ผู้โดยสารท่านใดยังไม่มีตั๋วใช้ในการโดยสาร สามารถติดต่อซื้อตั๋วโดยสารได้ที่ช่องจำหน่ายตั๋ว ขอบคุณครับ"
    elif idx == 1:
        text = f"โปรดทราบ ผู้โดยสารที่มีตั๋วใช้ในการโดยสารกับขบวนรถ ขบวนที่ {t_num} รับส่งผู้โดยสารต้นทาง {station(origin)} ปลายทาง {station(dest)} เที่ยวกำหนดเวลา {t_time} ขอให้ผู้โดยสารนำสิ่งของและสัมภาระของท่าน ไปรอรับการโดยสารในชานชาลาที่ {platform} ขอบคุณครับ"
    elif idx == 2:
        text = f"โปรดทราบ อีกสักครู่ ขบวนรถ ขบวนที่ {t_num} รับส่งผู้โดยสารต้นทาง {station(origin)} ปลายทาง {station(dest)} เที่ยวกำหนดเวลา {t_time} กำลังจะเข้าเทียบสถานีในชานชาลาที่ {platform} เพื่อความปลอดภัย กรุณายืนหลังเส้นสีเหลืองขอบชานชาลา และไม่เดินข้ามไปมา ระหว่างชานชาลาที่ {platform} ขอบคุณครับ"
    elif idx == 3:
        text = f"โปรดทราบ อีกสักครู่จะมีขบวนรถวิ่งผ่านสถานี บริเวณชานชาลาที่ {pass_platform} เพื่อความปลอดภัย กรุณายืนหลังเส้นสีเหลืองขอบชานชาลา และไม่เดินข้ามไปมา ระหว่างชานชาลาที่ {pass_platform} ขอบคุณครับ"
    elif idx == 4:
        text = f"โปรดทราบ ที่นี่{station(current)} ที่นี่{station(current)} ผู้โดยสารก่อนลงจากขบวนรถ โปรดตรวจสอบสิ่งของและสัมภาระของท่าน นำลงจากขบวนรถให้ครบถ้วน ขบวนรถที่จอดเทียบในชานชาลาที่ {platform} เป็นขบวนรถ ขบวนที่ {t_num} รับส่งผู้โดยสารต้นทาง {station(origin)} ปลายทาง {station(dest)} ขบวนรถเที่ยวนี้เมื่อออกจาก{station(current)} แล้ว จะหยุดรับส่งผู้โดยสารที่ {next_st} เป็นสถานีต่อไปตามลำดับ ขอบคุณครับ"
    elif idx == 5:
        text = f"โปรดทราบ วันนี้ขบวนรถ ขบวนที่ {t_num} รับส่งผู้โดยสารต้นทาง {station(origin)} ปลายทาง {station(dest)} เที่ยวกำหนดเวลา {t_time} ล่าช้ากว่ากำหนดเวลาเดิม คาดว่าจะถึง{station(current)} ได้ในเวลาโดยประมาณ {delay} ในนามของการรถไฟแห่งประเทศไทย ต้องขออภัยในความไม่สะดวกในครั้งนี้ ขอบคุณครับ"
    elif idx == 6:
        text = f"โปรดทราบ อีกสักครู่จะมีขบวนรถ ขบวนที่ {t_num} เข้าเทียบในชานชาลาที่ {platform} ผู้โดยสารที่ลงจากขบวนรถ โปรดระมัดระวังด้วยครับ ขอบคุณครับ"
    elif idx == 7:
        text = "ท่านผู้โดยสารโปรดทราบ เพื่อความปลอดภัยและสุขอนามัยที่ดี การรถไฟฯ ขอแจ้งให้ทราบว่า บริเวณสถานี บนขบวนรถ และภายในเขตพื้นที่สถานีทุกแห่ง เป็นเขตปลอดบุหรี่และเครื่องดื่มแอลกอฮอล์ ห้ามสูบบุหรี่และห้ามดื่มสุราโดยเด็ดขาด ผู้ฝ่าฝืนมีความผิดตามกฎหมาย ขอขอบคุณในความร่วมมือครับ"
    elif idx == 8:
        text = custom_text if custom_text.strip() else "กรุณาพิมพ์ข้อความที่ต้องการประกาศในช่องข้อความประกาศเองก่อนกดปุ่มครับ"
    elif idx == 9:
        text = f"โปรดทราบ อีกสักครู่จะมีขบวนรถ{train_type}วิ่งผ่านสถานี บริเวณชานชาลาที่ {pass_platform} เพื่อความปลอดภัย กรุณายืนหลังเส้นสีเหลืองขอบชานชาลา และไม่เดินข้ามไปมา ระหว่างชานชาลาที่ {pass_platform} ขอบคุณครับ"
    elif idx == 10:
        # ใช้แพตเทิร์นเดิมของประกาศรถเข้าพร้อมกัน และต่อประโยคแบบเดียวกันเมื่อมีขบวนที่ 3
        text = (
            f"ผู้โดยสารโปรดทราบ ที่นี่{station(current)} ที่นี่{station(current)} ก่อนผู้โดยสารจะลงจากขบวนรถ โปรดตรวจสอบสิ่งของและสัมภาระของท่าน นำลงให้ถูกต้องครบถ้วน "
            f"ขบวนรถที่จอดในชานชาลาที่ {platform} เป็นขบวนรถ ขบวนที่ {t_num} รับส่งผู้โดยสารต้นทาง {station(origin)} ปลายทาง {station(dest)} "
            f"และขบวนรถที่จอดในชานชาลาที่ {platform_2} เป็นขบวนรถ ขบวนที่ {t_num_2} รับส่งผู้โดยสารต้นทาง {station(origin_2)} ปลายทาง {station(dest_2)} "
        )
        if t_num_3:
            text += f"และขบวนรถที่จอดในชานชาลาที่ {platform_3} เป็นขบวนรถ ขบวนที่ {t_num_3} รับส่งผู้โดยสารต้นทาง {station(origin_3)} ปลายทาง {station(dest_3)} "
        text += (
            f"ขบวนรถในชานชาลาที่ {platform} เมื่อออกจาก{station(current)}แล้ว จะหยุดรับส่งผู้โดยสารที่ {next_st} เป็นสถานีต่อไปตามลำดับ "
            f"และขบวนรถในชานชาลาที่ {platform_2} เมื่อออกจาก{station(current)}แล้ว จะหยุดรับส่งผู้โดยสารที่ {next_st_2} เป็นสถานีต่อไปตามลำดับ "
        )
        if t_num_3:
            text += f"และขบวนรถในชานชาลาที่ {platform_3} เมื่อออกจาก{station(current)}แล้ว จะหยุดรับส่งผู้โดยสารที่ {next_st_3} เป็นสถานีต่อไปตามลำดับ "
        text += "ขอบคุณครับ"
    else:
        raise ValueError("ไม่พบประเภทประกาศที่เลือก")

    return clean_space(text)


LOGIN_HTML = r"""
<!doctype html><html lang="th"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>เข้าสู่ระบบ</title><style>
:root{--m:#800000;--md:#5b0000;--line:#eadfce;--muted:#766969}*{box-sizing:border-box}body{margin:0;min-height:100vh;display:grid;place-items:center;padding:18px;font-family:"Sarabun","Noto Sans Thai","Segoe UI",sans-serif;background:linear-gradient(145deg,#fffaf1,#f2e6d7);color:#251d1d}.box{width:min(440px,100%);background:#fff;border:1px solid var(--line);border-radius:24px;padding:26px;box-shadow:0 18px 50px rgba(80,20,20,.13)}h1{margin:0;color:var(--md)}p{color:var(--muted);line-height:1.6}label{display:block;margin:14px 0 6px;font-weight:800}input{width:100%;padding:13px;border:1px solid #d8c9b7;border-radius:13px;font:inherit}button{width:100%;margin-top:18px;padding:14px;border:0;border-radius:14px;background:linear-gradient(135deg,var(--md),var(--m));color:#fff;font:inherit;font-weight:900;cursor:pointer}.flash{padding:11px;border-radius:12px;background:#fff0f0;color:#9b1c1c}.note{padding:11px;border-radius:12px;background:#fff8dc;font-size:13px}
</style></head><body><form class="box" method="post"><h1>🚆 ระบบประกาศสถานี</h1><p>เข้าสู่ระบบเพื่อใช้งานและบันทึกประวัติผู้ประกาศ</p>{% for message in get_flashed_messages() %}<div class="flash">{{ message }}</div>{% endfor %}{% if show_default %}<div class="note"><b>บัญชีเริ่มต้น:</b> admin / admin1234<br>ควรเปลี่ยนรหัสผ่านทันทีหลังเข้าสู่ระบบ</div>{% endif %}<input type="hidden" name="next" value="{{ next_url }}"><label>ชื่อผู้ใช้</label><input name="username" autocomplete="username" required autofocus><label>รหัสผ่าน</label><input type="password" name="password" autocomplete="current-password" required><button>เข้าสู่ระบบ</button></form></body></html>
"""

MANAGEMENT_CSS = r"""
:root{--m:#800000;--md:#5b0000;--gold:#c7a12a;--bg:#f6eee3;--line:#e8dac6;--muted:#766969;--green:#177744;--red:#b42318}*{box-sizing:border-box}body{margin:0;padding:14px;font-family:"Sarabun","Noto Sans Thai","Segoe UI",sans-serif;background:linear-gradient(145deg,#fffaf1,var(--bg));color:#251d1d}a{color:var(--m)}.wrap{max-width:1320px;margin:auto}.head{padding:18px 20px;border-radius:20px;color:#fff;background:linear-gradient(135deg,var(--md),var(--m));box-shadow:0 10px 30px rgba(80,20,20,.12)}.head h1{margin:0;font-size:clamp(22px,3vw,31px)}.head p{margin:5px 0 0;opacity:.88}.nav{display:flex;flex-wrap:wrap;gap:8px;margin:11px 0}.nav a{padding:9px 12px;border-radius:11px;background:#fff;color:var(--md);text-decoration:none;font-weight:850;border:1px solid var(--line)}.nav a.active,.nav a:hover{background:var(--m);color:#fff}.grid{display:grid;grid-template-columns:minmax(320px,.68fr) minmax(0,1.5fr);gap:14px}.card{background:#fff;border:1px solid var(--line);border-radius:18px;overflow:hidden;box-shadow:0 10px 30px rgba(80,20,20,.08);margin-bottom:14px}.card h2{margin:0;padding:14px 17px;color:var(--md);font-size:18px;background:#fffdf9;border-bottom:1px solid var(--line)}.body{padding:16px}.fields{display:grid;grid-template-columns:1fr 1fr;gap:10px}.full{grid-column:1/-1}label{display:block;margin:0 0 5px;font-weight:800}input,select,textarea{width:100%;padding:11px 12px;border:1px solid #d8c9b7;border-radius:12px;background:#fff;font:inherit}textarea{min-height:75px;resize:vertical}.btn{display:inline-flex;align-items:center;justify-content:center;padding:10px 13px;border:0;border-radius:11px;background:var(--m);color:#fff;text-decoration:none;font:inherit;font-weight:850;cursor:pointer}.btn.gray{background:#665b55}.btn.green{background:var(--green)}.btn.red{background:var(--red)}.btn.gold{background:#a86900}.actions{display:flex;flex-wrap:wrap;gap:7px;margin-top:11px}.flash{padding:11px 13px;margin:10px 0;border-radius:12px;background:#fff3cd;color:#684f00;border:1px solid #ead28a}.table-wrap{overflow:auto}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:9px 10px;border-bottom:1px solid #eee4d6;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#fff9ef;color:var(--md);z-index:1}.muted{color:var(--muted);font-size:12px}.badge{display:inline-block;padding:4px 8px;border-radius:999px;font-size:11px;font-weight:850;background:#eee}.ok{background:#e8f6ed;color:#116735}.off{background:#fdebea;color:#a31d16}.warn{background:#fff3cd;color:#785b00}.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.stat{padding:14px;border-radius:14px;background:#fffaf1;border:1px solid var(--line)}.stat b{display:block;font-size:24px;color:var(--m)}details summary{cursor:pointer;color:var(--m);font-weight:850}.message{white-space:normal;min-width:280px;max-width:520px;line-height:1.5}.health-list{display:grid;gap:10px}.health-item{display:grid;grid-template-columns:42px 1fr;gap:11px;padding:13px;border:1px solid var(--line);border-radius:14px}.health-icon{width:42px;height:42px;display:grid;place-items:center;border-radius:12px;font-size:20px;background:#eee}@media(max-width:850px){.grid{grid-template-columns:1fr}.fields{grid-template-columns:1fr}.stats{grid-template-columns:1fr}body{padding:8px}}
"""

ADMIN_SCHEDULE_HTML = r"""
<!doctype html><html lang="th"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>จัดการตารางรถ</title><style>{{ management_css }}</style></head><body><main class="wrap"><header class="head"><h1>🚆 จัดการตารางรถ</h1><p>แก้ไขข้อมูลได้โดยไม่ต้องเปิดโค้ด · ตารางเก่ายังคงเก็บย้อนหลัง</p></header><nav class="nav"><a href="{{ url_for('index') }}">📢 หน้าประกาศ</a><a class="active" href="{{ url_for('admin_schedules') }}">🚆 ตารางรถ</a><a href="{{ url_for('admin_users') }}">👥 บัญชีผู้ใช้</a><a href="{{ url_for('history_page') }}">🕘 ประวัติ</a><a href="{{ url_for('health_page') }}">🩺 ตรวจสุขภาพ</a><a href="{{ url_for('logout') }}">ออกจากระบบ</a></nav>{% for message in get_flashed_messages() %}<div class="flash">{{ message }}</div>{% endfor %}
<div class="grid"><section>
<div class="card"><h2>สร้างตารางใหม่</h2><div class="body"><form method="post" action="{{ url_for('create_schedule_version') }}"><div class="fields"><div class="full"><label>ชื่อตาราง</label><input name="name" placeholder="เช่น ตารางเดินรถเดือนตุลาคม 2569" required></div><div><label>วันที่เริ่มใช้</label><input type="date" name="effective_date" value="{{ today }}" required></div><div><label>สถานะ</label><select name="status"><option value="draft">ฉบับร่าง</option><option value="published">ประกาศใช้</option></select></div><div class="full"><label>คัดลอกจากตารางเดิม</label><select name="copy_from"><option value="">เริ่มตารางว่าง</option>{% for v in versions %}<option value="{{ v.id }}">{{ v.name }} ({{ v.effective_date }})</option>{% endfor %}</select></div></div><div class="actions"><button class="btn">＋ สร้างตาราง</button></div></form></div></div>
<div class="card"><h2>ตารางย้อนหลัง</h2><div class="body"><form method="get"><label>เลือกตาราง</label><select name="version" onchange="this.form.submit()">{% for v in versions %}<option value="{{ v.id }}" {% if selected_version and v.id==selected_version.id %}selected{% endif %}>{{ v.effective_date }} · {{ v.name }} · {{ 'ใช้งาน' if v.status=='published' else 'ร่าง' }}</option>{% endfor %}</select></form>{% if selected_version %}<form method="post" action="{{ url_for('update_schedule_version', version_id=selected_version.id) }}" style="margin-top:12px"><div class="fields"><div class="full"><label>ชื่อ</label><input name="name" value="{{ selected_version.name }}" required></div><div><label>วันที่เริ่มใช้</label><input type="date" name="effective_date" value="{{ selected_version.effective_date }}" required></div><div><label>สถานะ</label><select name="status"><option value="draft" {% if selected_version.status=='draft' %}selected{% endif %}>ฉบับร่าง</option><option value="published" {% if selected_version.status=='published' %}selected{% endif %}>ประกาศใช้</option></select></div></div><div class="actions"><button class="btn gold">บันทึกข้อมูลตาราง</button></div></form>{% endif %}</div></div>
<div class="card"><h2>นำเข้า / สำรอง / กู้คืน</h2><div class="body"><form method="post" enctype="multipart/form-data" action="{{ url_for('import_schedule') }}"><input type="hidden" name="version_id" value="{{ selected_version.id if selected_version else '' }}"><label>นำเข้า Excel หรือ CSV</label><input type="file" name="file" accept=".xlsx,.csv" required><p class="muted">หัวคอลัมน์: direction, num, origin, dest, time, next, service_pattern, service_dates, enabled</p><button class="btn green">นำเข้าข้อมูล</button></form><hr style="border:0;border-top:1px solid var(--line);margin:16px 0"><div class="actions"><a class="btn gray" href="{{ url_for('backup_data') }}">⬇ สำรองข้อมูล JSON</a></div><form method="post" enctype="multipart/form-data" action="{{ url_for('restore_data') }}" style="margin-top:12px" onsubmit="return confirm('ยืนยันกู้คืนข้อมูล? ตารางและประวัติปัจจุบันจะถูกแทนที่')"><label>กู้คืนจากไฟล์ JSON</label><input type="file" name="file" accept=".json" required><p class="muted">เพื่อป้องกันล็อกอินไม่ได้ ระบบจะไม่เขียนทับบัญชีผู้ใช้</p><button class="btn red">กู้คืนข้อมูล</button></form></div></div>
</section><section>
<div class="card"><h2>{{ 'แก้ไขขบวน '+edit_train.num if edit_train else 'เพิ่มขบวนรถ' }}</h2><div class="body"><form method="post" action="{{ url_for('save_train') }}"><input type="hidden" name="version_id" value="{{ selected_version.id if selected_version else '' }}"><input type="hidden" name="train_id" value="{{ edit_train.id if edit_train else '' }}"><div class="fields"><div><label>ทิศทาง</label><select name="direction"><option value="inbound" {% if edit_train and edit_train.direction=='inbound' %}selected{% endif %}>ขาเข้า กรุงเทพ</option><option value="outbound" {% if edit_train and edit_train.direction=='outbound' %}selected{% endif %}>ขาออก ไปทางตะวันออก</option></select></div><div><label>เลขขบวน</label><input name="num" value="{{ edit_train.num if edit_train else '' }}" required></div><div><label>ต้นทาง</label><input name="origin" value="{{ edit_train.origin if edit_train else '' }}" required></div><div><label>ปลายทาง</label><input name="dest" value="{{ edit_train.dest if edit_train else '' }}" required></div><div><label>เวลา HH:MM</label><input type="time" name="time_hhmm" value="{{ edit_train.time_hhmm if edit_train else '' }}" required></div><div><label>วันให้บริการ</label><select name="service_pattern">{% for key,label in service_labels.items() %}<option value="{{ key }}" {% if edit_train and edit_train.service_pattern==key %}selected{% endif %}>{{ label }}</option>{% endfor %}</select></div><div class="full"><label>สถานีต่อไป</label><input name="next_station" value="{{ edit_train.next_station if edit_train else '' }}"></div><div class="full"><label>วันที่ให้บริการเฉพาะ (YYYY-MM-DD คั่นด้วย comma)</label><textarea name="service_dates" placeholder="2026-07-28, 2026-08-12">{{ edit_train.service_dates if edit_train else '' }}</textarea></div><div class="full"><label><input style="width:auto" type="checkbox" name="enabled" value="1" {% if not edit_train or edit_train.enabled %}checked{% endif %}> เปิดใช้งานขบวนนี้</label></div></div><div class="actions"><button class="btn">บันทึกขบวน</button>{% if edit_train %}<a class="btn gray" href="{{ url_for('admin_schedules',version=selected_version.id) }}">ยกเลิกแก้ไข</a>{% endif %}</div></form></div></div>
<div class="card"><h2>รายการขบวนในตาราง{% if selected_version %} · {{ selected_version.name }}{% endif %}</h2><div class="body"><div class="stats"><div class="stat"><span>ทั้งหมด</span><b>{{ trains|length }}</b></div><div class="stat"><span>เปิดใช้งาน</span><b>{{ trains|selectattr('enabled')|list|length }}</b></div><div class="stat"><span>วันที่เริ่มใช้</span><b style="font-size:17px">{{ selected_version.effective_date if selected_version else '-' }}</b></div></div></div><div class="table-wrap"><table><thead><tr><th>สถานะ</th><th>ขบวน</th><th>เวลา</th><th>เส้นทาง</th><th>วันให้บริการ</th><th>สถานีต่อไป</th><th>จัดการ</th></tr></thead><tbody>{% for t in trains %}<tr><td><span class="badge {{ 'ok' if t.enabled else 'off' }}">{{ 'เปิด' if t.enabled else 'ปิด' }}</span></td><td><b>{{ t.num }}</b><div class="muted">{{ direction_labels[t.direction] }}</div></td><td>{{ t.time_hhmm }}</td><td>{{ t.origin }} → {{ t.dest }}</td><td>{{ service_labels[t.service_pattern] }}{% if t.service_dates %}<div class="muted">{{ t.service_dates }}</div>{% endif %}</td><td>{{ t.next_station }}</td><td><div class="actions" style="margin:0"><a class="btn gold" href="{{ url_for('admin_schedules',version=selected_version.id,edit=t.id) }}">แก้ไข</a><form method="post" action="{{ url_for('toggle_train',train_id=t.id) }}"><input type="hidden" name="version_id" value="{{ selected_version.id }}"><button class="btn gray">{{ 'ปิดชั่วคราว' if t.enabled else 'เปิดขบวน' }}</button></form><form method="post" action="{{ url_for('delete_train',train_id=t.id) }}" onsubmit="return confirm('ลบขบวนนี้หรือไม่?')"><input type="hidden" name="version_id" value="{{ selected_version.id }}"><button class="btn red">ลบ</button></form></div></td></tr>{% else %}<tr><td colspan="7">ยังไม่มีขบวนรถในตารางนี้</td></tr>{% endfor %}</tbody></table></div></div>
</section></div></main></body></html>
"""

USERS_HTML = r"""
<!doctype html><html lang="th"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>บัญชีผู้ใช้</title><style>{{ management_css }}</style></head><body><main class="wrap"><header class="head"><h1>👥 บัญชีผู้ใช้และสิทธิ์</h1><p>แยกผู้ประกาศ ผู้ดูแลระบบ และผู้ตรวจสอบประวัติ</p></header><nav class="nav"><a href="{{ url_for('index') }}">📢 หน้าประกาศ</a><a href="{{ url_for('admin_schedules') }}">🚆 ตารางรถ</a><a class="active" href="{{ url_for('admin_users') }}">👥 บัญชีผู้ใช้</a><a href="{{ url_for('history_page') }}">🕘 ประวัติ</a><a href="{{ url_for('health_page') }}">🩺 ตรวจสุขภาพ</a><a href="{{ url_for('logout') }}">ออกจากระบบ</a></nav>{% for message in get_flashed_messages() %}<div class="flash">{{ message }}</div>{% endfor %}<div class="grid"><section><div class="card"><h2>เพิ่มบัญชีใหม่</h2><div class="body"><form method="post" action="{{ url_for('create_user') }}"><label>ชื่อผู้ใช้</label><input name="username" required><label style="margin-top:10px">ชื่อที่แสดง</label><input name="display_name" required><label style="margin-top:10px">สิทธิ์</label><select name="role">{% for key,label in role_labels.items() %}<option value="{{ key }}">{{ label }}</option>{% endfor %}</select><label style="margin-top:10px">รหัสผ่านเริ่มต้น</label><input type="password" name="password" minlength="6" required><div class="actions"><button class="btn">เพิ่มบัญชี</button></div></form></div></div></section><section><div class="card"><h2>บัญชีทั้งหมด</h2><div class="table-wrap"><table><thead><tr><th>ผู้ใช้</th><th>สิทธิ์</th><th>สถานะ</th><th>เข้าใช้ล่าสุด</th><th>แก้ไข</th></tr></thead><tbody>{% for u in users %}<tr><td><b>{{ u.display_name }}</b><div class="muted">{{ u.username }}</div></td><td>{{ role_labels[u.role] }}</td><td><span class="badge {{ 'ok' if u.active else 'off' }}">{{ 'ใช้งาน' if u.active else 'ปิด' }}</span></td><td>{{ u.last_login or '-' }}</td><td><form method="post" action="{{ url_for('update_user',user_id=u.id) }}"><div class="fields" style="min-width:460px"><div><input name="display_name" value="{{ u.display_name }}" required></div><div><select name="role">{% for key,label in role_labels.items() %}<option value="{{ key }}" {% if u.role==key %}selected{% endif %}>{{ label }}</option>{% endfor %}</select></div><div><input type="password" name="password" placeholder="รหัสใหม่ (เว้นว่างได้)" minlength="6"></div><div><label><input style="width:auto" type="checkbox" name="active" value="1" {% if u.active %}checked{% endif %}> เปิดใช้งาน</label></div></div><button class="btn gold" style="margin-top:7px">บันทึก</button></form></td></tr>{% endfor %}</tbody></table></div></div></section></div></main></body></html>
"""

HISTORY_HTML = r"""
<!doctype html><html lang="th"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>ประวัติการประกาศ</title><style>{{ management_css }}</style></head><body><main class="wrap"><header class="head"><h1>🕘 ประวัติการประกาศ</h1><p>ใช้ตรวจหาสาเหตุเมื่อระบบมีปัญหา ไม่ใช่เพื่อจับผิดผู้ใช้งาน</p></header><nav class="nav"><a href="{{ url_for('index') }}">📢 หน้าประกาศ</a>{% if current_user.role=='admin' %}<a href="{{ url_for('admin_schedules') }}">🚆 ตารางรถ</a><a href="{{ url_for('admin_users') }}">👥 บัญชีผู้ใช้</a>{% endif %}<a class="active" href="{{ url_for('history_page') }}">🕘 ประวัติ</a><a href="{{ url_for('health_page') }}">🩺 ตรวจสุขภาพ</a><a href="{{ url_for('logout') }}">ออกจากระบบ</a></nav><div class="card"><h2>ตัวกรอง</h2><div class="body"><form method="get"><div class="fields"><div><label>ตั้งแต่วันที่</label><input type="date" name="date_from" value="{{ filters.date_from }}"></div><div><label>ถึงวันที่</label><input type="date" name="date_to" value="{{ filters.date_to }}"></div><div><label>เลขขบวน</label><input name="train_num" value="{{ filters.train_num }}"></div><div><label>ผู้ประกาศ</label><input name="username" value="{{ filters.username }}"></div></div><div class="actions"><button class="btn">ค้นหา</button><a class="btn gray" href="{{ url_for('history_page') }}">ล้างตัวกรอง</a></div></form></div></div><div class="card"><h2>พบ {{ histories|length }} รายการล่าสุด</h2><div class="table-wrap"><table><thead><tr><th>วันเวลา</th><th>ผู้ประกาศ</th><th>ขบวน/ประเภท</th><th>ภาษา/เสียง</th><th>ชานชาลา</th><th>สร้างเสียง</th><th>ผลการเล่น</th><th>Pause / Stop</th><th>ข้อความ</th></tr></thead><tbody>{% for h in histories %}<tr><td>{{ h.started_at }}</td><td>{{ h.username }}</td><td><b>{{ h.train_num or '-' }}</b><div class="muted">{{ h.announcement_type }}</div></td><td>{{ h.announce_mode }}<div class="muted">{{ h.voice }}</div></td><td>{{ h.platform or '-' }}</td><td>{{ h.generation_ms if h.generation_ms is not none else '-' }} ms</td><td>{% if h.playback_success==1 %}<span class="badge ok">สำเร็จ</span>{% elif h.playback_success==0 %}<span class="badge off">ไม่สำเร็จ/หยุด</span>{% else %}<span class="badge warn">ยังไม่จบ</span>{% endif %}{% if h.failure_reason %}<div class="muted">{{ h.failure_reason }}</div>{% endif %}</td><td>พัก {{ h.pause_count }} ครั้ง<div class="muted">Stop: {{ h.stop_time or '-' }}</div></td><td class="message"><details><summary>ดูข้อความและเหตุการณ์</summary><div style="margin:8px 0">{{ h.message|safe if h.message else '-' }}</div><a href="{{ url_for('history_detail',history_id=h.id) }}">ดูเหตุการณ์ทั้งหมด</a></details></td></tr>{% else %}<tr><td colspan="9">ยังไม่มีประวัติ</td></tr>{% endfor %}</tbody></table></div></div></main></body></html>
"""

HISTORY_DETAIL_HTML = r"""
<!doctype html><html lang="th"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>รายละเอียดประวัติ</title><style>{{ management_css }}</style></head><body><main class="wrap"><header class="head"><h1>รายละเอียดประวัติ #{{ history.id }}</h1><p>{{ history.started_at }} · {{ history.username }} · ขบวน {{ history.train_num or '-' }}</p></header><nav class="nav"><a href="{{ url_for('history_page') }}">← กลับหน้าประวัติ</a><a href="{{ url_for('index') }}">หน้าประกาศ</a></nav><div class="card"><h2>ข้อความประกาศ</h2><div class="body message">{{ history.message|safe if history.message else '-' }}</div></div><div class="card"><h2>ลำดับเหตุการณ์</h2><div class="table-wrap"><table><thead><tr><th>เวลา</th><th>เหตุการณ์</th><th>รายละเอียด</th></tr></thead><tbody>{% for e in events %}<tr><td>{{ e.event_at }}</td><td>{{ e.event_type }}</td><td class="message">{{ e.details or '-' }}</td></tr>{% endfor %}</tbody></table></div></div></main></body></html>
"""

HEALTH_HTML = r"""
<!doctype html><html lang="th"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>ตรวจสุขภาพระบบ</title><style>{{ management_css }}</style></head><body><main class="wrap"><header class="head"><h1>🩺 ตรวจสุขภาพระบบ</h1><p>ตรวจ Backend, TTS, ไฟล์เสียง, พื้นที่จัดเก็บ, เวลาเครื่อง และอุปกรณ์เสียง</p></header><nav class="nav"><a href="{{ url_for('index') }}">📢 หน้าประกาศ</a>{% if current_user.role=='admin' %}<a href="{{ url_for('admin_schedules') }}">🚆 ตารางรถ</a><a href="{{ url_for('admin_users') }}">👥 บัญชีผู้ใช้</a>{% endif %}{% if current_user.role in ['admin','auditor'] %}<a href="{{ url_for('history_page') }}">🕘 ประวัติ</a>{% endif %}<a class="active" href="{{ url_for('health_page') }}">🩺 ตรวจสุขภาพ</a><a href="{{ url_for('logout') }}">ออกจากระบบ</a></nav><div class="card"><h2>ผลตรวจอัตโนมัติ</h2><div class="body"><div id="healthList" class="health-list"><div class="health-item"><div class="health-icon">⏳</div><div><b>กำลังตรวจสอบ...</b></div></div></div><div class="actions"><button class="btn" onclick="runHealth()">ตรวจใหม่</button><button class="btn gold" onclick="testSpeaker()">▶ ทดลองลำโพงด้วยเสียงเตือน</button></div><p class="muted">เบราว์เซอร์ไม่สามารถรับรองว่าลำโพงภายนอกเปิดอยู่จริงได้ จึงมีปุ่มทดลองเสียงสำหรับยืนยันด้วยการฟัง</p></div></div></main><script>
function esc(v){const d=document.createElement('div');d.textContent=v??'';return d.innerHTML}function item(label,ok,detail){return `<div class="health-item"><div class="health-icon" style="background:${ok?'#e8f6ed':'#fdebea'}">${ok?'✓':'!'}</div><div><b>${esc(label)}</b><div class="muted">${esc(detail)}</div></div></div>`}async function runHealth(){const list=document.getElementById('healthList');list.innerHTML=item('กำลังตรวจสอบ',true,'กรุณารอสักครู่');try{const started=Date.now();const r=await fetch('/api/health',{cache:'no-store'});const data=await r.json();let html=data.checks.map(c=>item(c.label,c.ok,c.detail)).join('');const delta=Math.abs(Date.now()-data.server_epoch_ms);html+=item('เวลาเครื่อง',delta<120000,`เวลาต่างจาก Backend ${Math.round(delta/1000)} วินาที`);let audioOk=!!(window.Audio&&document.createElement('audio').canPlayType('audio/mpeg'));let detail=audioOk?'เบราว์เซอร์รองรับ MP3 และระบบเสียง':'เบราว์เซอร์ไม่รองรับการเล่น MP3';try{if(navigator.mediaDevices?.enumerateDevices){const devices=await navigator.mediaDevices.enumerateDevices();const outputs=devices.filter(d=>d.kind==='audiooutput');if(outputs.length)detail+=` · พบช่องเสียงออก ${outputs.length} รายการ`;}}catch(e){}html+=item('ลำโพง / อุปกรณ์เสียง',audioOk,detail);list.innerHTML=html}catch(e){list.innerHTML=item('Backend',false,e.message)}}async function testSpeaker(){try{const a=new Audio('/audio/chime.mp3');await a.play()}catch(e){alert('เล่นเสียงไม่ได้: '+e.message)}}runHealth();
</script></body></html>
"""


@app.route("/login", methods=["GET", "POST"])
def login():
    if get_current_user():
        return redirect(url_for("index"))
    next_url = request.values.get("next") or url_for("index")
    if not next_url.startswith("/") or next_url.startswith("//"):
        next_url = url_for("index")
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        with get_db() as conn:
            user = conn.execute("SELECT * FROM users WHERE username=? COLLATE NOCASE", (username,)).fetchone()
            if user and user["active"] and check_password_hash(user["password_hash"], password):
                session.clear()
                _store_session_user(user)
                conn.execute("UPDATE users SET last_login=? WHERE id=?", (now_iso(), user["id"]))
                destination = next_url
                if user["role"] == "auditor" and destination == url_for("index"):
                    destination = url_for("history_page")
                return redirect(destination)
        flash("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง")
    with get_db() as conn:
        show_default = (not os.environ.get("INITIAL_ADMIN_PASSWORD") and
                        conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 1 and
                        not conn.execute("SELECT last_login FROM users LIMIT 1").fetchone()[0])
    return render_template_string(LOGIN_HTML, next_url=next_url, show_default=show_default)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin/schedules")
@roles_required("admin")
def admin_schedules():
    with get_db() as conn:
        versions = [dict(row) for row in conn.execute("SELECT * FROM schedule_versions ORDER BY effective_date DESC,id DESC").fetchall()]
        version_id = request.args.get("version", type=int) or (versions[0]["id"] if versions else None)
        selected = next((v for v in versions if v["id"] == version_id), None)
        trains = [dict(row) for row in conn.execute("SELECT * FROM trains WHERE version_id=? ORDER BY direction,time_hhmm,num", (version_id,)).fetchall()] if version_id else []
        edit_id = request.args.get("edit", type=int)
        edit_row = conn.execute("SELECT * FROM trains WHERE id=? AND version_id=?", (edit_id, version_id)).fetchone() if edit_id and version_id else None
        edit_train = dict(edit_row) if edit_row else None
    return render_template_string(ADMIN_SCHEDULE_HTML, management_css=MANAGEMENT_CSS, versions=versions,
                                  selected_version=selected, trains=trains, edit_train=edit_train,
                                  service_labels=SERVICE_LABELS, direction_labels=DIRECTION_LABELS,
                                  today=now_bangkok().date().isoformat())


@app.route("/admin/schedules/version/create", methods=["POST"])
@roles_required("admin")
def create_schedule_version():
    name = (request.form.get("name") or "").strip()
    effective_date = (request.form.get("effective_date") or "").strip()
    status = request.form.get("status") or "draft"
    copy_from = request.form.get("copy_from", type=int)
    try:
        date.fromisoformat(effective_date)
        if not name or status not in {"draft", "published"}:
            raise ValueError("ข้อมูลตารางไม่ครบ")
        user = get_current_user()
        with get_db() as conn:
            cursor = conn.execute("INSERT INTO schedule_versions(name,effective_date,status,created_at,created_by) VALUES(?,?,?,?,?)", (name,effective_date,status,now_iso(),user["id"]))
            new_id = cursor.lastrowid
            if copy_from:
                conn.execute("""INSERT INTO trains(version_id,direction,num,label,origin,dest,time_hhmm,time_spoken,next_station,service_pattern,service_dates,enabled,created_at,updated_at)
                              SELECT ?,direction,num,label,origin,dest,time_hhmm,time_spoken,next_station,service_pattern,service_dates,enabled,?,? FROM trains WHERE version_id=?""", (new_id,now_iso(),now_iso(),copy_from))
        invalidate_train_cache()
        flash("สร้างตารางใหม่แล้ว")
        return redirect(url_for("admin_schedules", version=new_id))
    except Exception as exc:
        flash(str(exc))
        return redirect(url_for("admin_schedules"))


@app.route("/admin/schedules/version/<int:version_id>/update", methods=["POST"])
@roles_required("admin")
def update_schedule_version(version_id):
    name = (request.form.get("name") or "").strip()
    effective_date = (request.form.get("effective_date") or "").strip()
    status = request.form.get("status") or "draft"
    try:
        date.fromisoformat(effective_date)
        if not name or status not in {"draft", "published"}: raise ValueError("ข้อมูลไม่ถูกต้อง")
        with get_db() as conn:
            conn.execute("UPDATE schedule_versions SET name=?,effective_date=?,status=? WHERE id=?", (name,effective_date,status,version_id))
        invalidate_train_cache()
        flash("บันทึกข้อมูลตารางแล้ว")
    except Exception as exc: flash(str(exc))
    return redirect(url_for("admin_schedules", version=version_id))


@app.route("/admin/trains/save", methods=["POST"])
@roles_required("admin")
def save_train():
    version_id = request.form.get("version_id", type=int)
    try:
        save_train_record(request.form)
        invalidate_train_cache()
        flash("บันทึกขบวนรถแล้ว และตรวจสอบข้อมูลซ้ำเรียบร้อย")
    except Exception as exc:
        flash(str(exc))
    return redirect(url_for("admin_schedules", version=version_id))


@app.route("/admin/trains/<int:train_id>/toggle", methods=["POST"])
@roles_required("admin")
def toggle_train(train_id):
    version_id = request.form.get("version_id", type=int)
    with get_db() as conn:
        conn.execute("UPDATE trains SET enabled=CASE enabled WHEN 1 THEN 0 ELSE 1 END,updated_at=? WHERE id=?", (now_iso(),train_id))
    invalidate_train_cache()
    flash("เปลี่ยนสถานะขบวนแล้ว")
    return redirect(url_for("admin_schedules", version=version_id))


@app.route("/admin/trains/<int:train_id>/delete", methods=["POST"])
@roles_required("admin")
def delete_train(train_id):
    version_id = request.form.get("version_id", type=int)
    with get_db() as conn: conn.execute("DELETE FROM trains WHERE id=?", (train_id,))
    invalidate_train_cache()
    flash("ลบขบวนแล้ว")
    return redirect(url_for("admin_schedules", version=version_id))


@app.route("/admin/schedules/import", methods=["POST"])
@roles_required("admin")
def import_schedule():
    version_id = request.form.get("version_id", type=int)
    file = request.files.get("file")
    try:
        if not version_id or not file: raise ValueError("กรุณาเลือกตารางและไฟล์")
        inserted, skipped = import_schedule_file(file, version_id)
        invalidate_train_cache()
        message = f"นำเข้าสำเร็จ {inserted} ขบวน"
        if skipped: message += f" · ข้าม {len(skipped)} แถว: " + " | ".join(skipped[:5])
        flash(message)
    except Exception as exc: flash(str(exc))
    return redirect(url_for("admin_schedules", version=version_id))


@app.route("/admin/backup")
@roles_required("admin")
def backup_data():
    tables = ["schedule_versions","trains","users","announcement_history","announcement_events"]
    payload = {"schema_version":1,"created_at":now_iso(),"tables":{}}
    with get_db() as conn:
        for table in tables:
            payload["tables"][table] = [dict(row) for row in conn.execute(f"SELECT * FROM {table}").fetchall()]
    stream = io.BytesIO(json.dumps(payload,ensure_ascii=False,indent=2).encode("utf-8"))
    filename = f"station_backup_{now_bangkok().strftime('%Y%m%d_%H%M%S')}.json"
    return send_file(stream,mimetype="application/json",as_attachment=True,download_name=filename)


@app.route("/admin/restore", methods=["POST"])
@roles_required("admin")
def restore_data():
    file = request.files.get("file")
    try:
        if not file: raise ValueError("ไม่พบไฟล์สำรอง")
        payload = json.load(file)
        tables = payload.get("tables") or {}
        versions = tables.get("schedule_versions",[]); trains=tables.get("trains",[])
        histories=tables.get("announcement_history",[]); events=tables.get("announcement_events",[])
        if not isinstance(versions,list) or not isinstance(trains,list): raise ValueError("รูปแบบไฟล์สำรองไม่ถูกต้อง")
        user=get_current_user()
        with get_db() as conn:
            conn.execute("DELETE FROM announcement_events"); conn.execute("DELETE FROM announcement_history"); conn.execute("DELETE FROM trains"); conn.execute("DELETE FROM schedule_versions")
            for v in versions:
                conn.execute("INSERT INTO schedule_versions(id,name,effective_date,status,created_at,created_by) VALUES(?,?,?,?,?,?)", (v.get('id'),v.get('name'),v.get('effective_date'),v.get('status','published'),v.get('created_at',now_iso()),user['id']))
            for t in trains:
                conn.execute("""INSERT INTO trains(id,version_id,direction,num,label,origin,dest,time_hhmm,time_spoken,next_station,service_pattern,service_dates,enabled,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (t.get('id'),t.get('version_id'),t.get('direction'),t.get('num'),t.get('label'),t.get('origin'),t.get('dest'),t.get('time_hhmm'),t.get('time_spoken'),t.get('next_station',''),t.get('service_pattern','daily'),t.get('service_dates',''),t.get('enabled',1),t.get('created_at',now_iso()),t.get('updated_at',now_iso())))
            for h in histories:
                conn.execute("""INSERT INTO announcement_history(id,started_at,user_id,username,train_num,announcement_type,announce_mode,voice,platform,message,generation_ms,playback_success,pause_times,stop_time,completed_at,failure_reason) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (h.get('id'),h.get('started_at'),None,h.get('username','ไม่ทราบ'),h.get('train_num'),h.get('announcement_type',''),h.get('announce_mode',''),h.get('voice',''),h.get('platform'),h.get('message'),h.get('generation_ms'),h.get('playback_success'),h.get('pause_times','[]'),h.get('stop_time'),h.get('completed_at'),h.get('failure_reason')))
            for e in events:
                conn.execute("INSERT INTO announcement_events(id,history_id,event_type,event_at,details) VALUES(?,?,?,?,?)", (e.get('id'),e.get('history_id'),e.get('event_type'),e.get('event_at'),e.get('details')))
            if USE_POSTGRES:
                for table in ("schedule_versions", "trains", "announcement_history", "announcement_events"):
                    conn.execute(
                        f"SELECT setval(pg_get_serial_sequence('{table}','id'), "
                        f"COALESCE((SELECT MAX(id) FROM {table}), 1), "
                        f"EXISTS(SELECT 1 FROM {table}))"
                    )
        invalidate_train_cache()
        flash("กู้คืนตารางรถและประวัติแล้ว บัญชีผู้ใช้เดิมยังคงอยู่")
    except Exception as exc: flash(f"กู้คืนไม่สำเร็จ: {exc}")
    return redirect(url_for("admin_schedules"))


@app.route("/admin/users")
@roles_required("admin")
def admin_users():
    with get_db() as conn: users=[dict(r) for r in conn.execute("SELECT * FROM users ORDER BY active DESC,role,username").fetchall()]
    return render_template_string(USERS_HTML,management_css=MANAGEMENT_CSS,users=users,role_labels=ROLE_LABELS)


@app.route("/admin/users/create",methods=["POST"])
@roles_required("admin")
def create_user():
    username=(request.form.get('username') or '').strip(); display=(request.form.get('display_name') or '').strip(); role=request.form.get('role'); password=request.form.get('password') or ''
    try:
        if not username or not display or role not in ROLE_LABELS or len(password)<6: raise ValueError("กรอกข้อมูลให้ครบ และรหัสผ่านอย่างน้อย 6 ตัว")
        with get_db() as conn: conn.execute("INSERT INTO users(username,password_hash,display_name,role,active,created_at) VALUES(?,?,?,?,1,?)",(username,generate_password_hash(password),display,role,now_iso()))
        flash("เพิ่มบัญชีแล้ว")
    except DB_INTEGRITY_ERRORS: flash("ชื่อผู้ใช้นี้มีอยู่แล้ว")
    except Exception as exc: flash(str(exc))
    return redirect(url_for('admin_users'))


@app.route("/admin/users/<int:user_id>/update",methods=["POST"])
@roles_required("admin")
def update_user(user_id):
    current=get_current_user(); display=(request.form.get('display_name') or '').strip(); role=request.form.get('role'); password=request.form.get('password') or ''; active=1 if request.form.get('active') else 0
    try:
        if not display or role not in ROLE_LABELS: raise ValueError("ข้อมูลไม่ถูกต้อง")
        if user_id==current['id'] and not active: raise ValueError("ไม่สามารถปิดบัญชีที่กำลังใช้งาน")
        with get_db() as conn:
            existing=conn.execute("SELECT role,active FROM users WHERE id=?",(user_id,)).fetchone()
            if not existing: raise ValueError("ไม่พบบัญชี")
            if existing['role']=='admin' and (role!='admin' or not active):
                count=conn.execute("SELECT COUNT(*) FROM users WHERE role='admin' AND active=1").fetchone()[0]
                if count<=1: raise ValueError("ต้องมีผู้ดูแลระบบที่เปิดใช้งานอย่างน้อย 1 บัญชี")
            conn.execute("UPDATE users SET display_name=?,role=?,active=? WHERE id=?",(display,role,active,user_id))
            if password:
                if len(password)<6: raise ValueError("รหัสผ่านอย่างน้อย 6 ตัว")
                conn.execute("UPDATE users SET password_hash=? WHERE id=?",(generate_password_hash(password),user_id))
        if user_id == current["id"]:
            _store_session_user({"id": user_id, "username": current["username"], "display_name": display, "role": role, "active": active})
        flash("บันทึกบัญชีแล้ว")
    except Exception as exc: flash(str(exc))
    return redirect(url_for('admin_users'))


@app.route("/history")
@roles_required("admin","auditor")
def history_page():
    filters={k:(request.args.get(k) or '').strip() for k in ('date_from','date_to','train_num','username')}
    where=[]; params=[]
    if filters['date_from']: where.append("substr(started_at,1,10)>=?"); params.append(filters['date_from'])
    if filters['date_to']: where.append("substr(started_at,1,10)<=?"); params.append(filters['date_to'])
    if filters['train_num']: where.append("train_num LIKE ?"); params.append('%'+filters['train_num']+'%')
    if filters['username']: where.append("username LIKE ?"); params.append('%'+filters['username']+'%')
    sql="SELECT * FROM announcement_history"+(" WHERE "+" AND ".join(where) if where else "")+" ORDER BY id DESC LIMIT 500"
    with get_db() as conn: rows=[dict(r) for r in conn.execute(sql,params).fetchall()]
    for row in rows:
        try: row['pause_count']=len(json.loads(row.get('pause_times') or '[]'))
        except Exception: row['pause_count']=0
    return render_template_string(HISTORY_HTML,management_css=MANAGEMENT_CSS,histories=rows,filters=filters,current_user=get_current_user())


@app.route("/history/<int:history_id>")
@roles_required("admin","auditor")
def history_detail(history_id):
    with get_db() as conn:
        history=conn.execute("SELECT * FROM announcement_history WHERE id=?",(history_id,)).fetchone()
        events=[dict(r) for r in conn.execute("SELECT * FROM announcement_events WHERE history_id=? ORDER BY id",(history_id,)).fetchall()]
    if not history: abort(404)
    return render_template_string(HISTORY_DETAIL_HTML,management_css=MANAGEMENT_CSS,history=dict(history),events=events)


@app.route("/health")
@login_required
def health_page():
    return render_template_string(HEALTH_HTML,management_css=MANAGEMENT_CSS,current_user=get_current_user())


@app.route("/api/health")
@login_required
def api_health():
    return jsonify(status="success",checks=backend_health_checks(),server_epoch_ms=int(time.time()*1000),server_time=now_iso())


@app.route("/api/history/start",methods=["POST"])
@roles_required("announcer", "admin")
def api_history_start():
    try:
        payload=request.get_json(silent=True) or {}
        history_id=insert_history(payload,get_current_user())
        return jsonify(status="success",history_id=history_id)
    except Exception as exc:
        return jsonify(status="error",message=str(exc)),500


@app.route("/api/history/event",methods=["POST"])
@roles_required("announcer", "admin")
def api_history_event():
    try:
        data=request.get_json(silent=True) or {}; history_id=int(data.get('history_id')); event_type=(data.get('event_type') or '').strip()
        if event_type not in {'generated','pause','resume','stop','success','failed'}: raise ValueError("ประเภทเหตุการณ์ไม่ถูกต้อง")
        event_at=add_history_event(history_id,event_type,data.get('details') or {})
        return jsonify(status="success",event_at=event_at)
    except Exception as exc:
        return jsonify(status="error",message=str(exc)),400


@app.route("/")
@roles_required("announcer", "admin")
def index():
    inbound, outbound, train_data, schedule_version = get_active_train_lists()
    user = get_current_user()
    return render_template_string(
        HTML_PAGE,
        inbound=inbound,
        outbound=outbound,
        trains_json=json.dumps(train_data, ensure_ascii=False),
        grouped_buttons=group_buttons(ANNOUNCEMENT_BUTTONS),
        voice_name=VOICE_NAME,
        en_voice_name=english_voice_for(VOICE_NAME),
        current_user=user,
        role_labels=ROLE_LABELS,
        schedule_version=schedule_version,
    )


@app.route("/audio/<path:filename>")
def serve_audio(filename):
    # ไฟล์เสียงเตือนเดิม ให้วาง chime.mp3 ไว้โฟลเดอร์เดียวกับไฟล์ Python
    if filename == CHIME_FILENAME:
        return send_from_directory(BASE_DIR, CHIME_FILENAME, mimetype="audio/mpeg", as_attachment=False, conditional=True, max_age=604800)

    # ไฟล์เสียงประกาศที่ระบบสร้างใหม่ จะอยู่ในโฟลเดอร์ audio_generated
    return send_from_directory(AUDIO_DIR, filename, mimetype="audio/mpeg", as_attachment=False, conditional=True, max_age=604800)


@app.route("/test-station-voice", methods=["POST"])
@roles_required("announcer", "admin")
def test_station_voice():
    cleanup_old_audio()
    data = request.get_json(silent=True) or {}
    thai_voice = (data.get("thai_voice") or VOICE_NAME).strip()
    if thai_voice not in THAI_VOICE_OPTIONS:
        thai_voice = VOICE_NAME

    mode = (data.get("announce_mode") or "thai_only").strip()
    if mode not in {"thai_only", "english_only", "bilingual"}:
        mode = "thai_only"

    english_voice = english_voice_for(thai_voice)
    segments = []
    thai_text = ""
    english_text = ""

    if mode in {"thai_only", "bilingual"}:
        thai_text = "ผู้โดยสารโปรดทราบ ที่นี่สถานีคลองบางพระ ที่นี่สถานีคลองบางพระ ขอบคุณครับ"
        thai_text = apply_voice_politeness(thai_text, thai_voice)
        segments.append({
            "code": "th_test",
            "label": "ภาษาไทย",
            "text": prepare_tts_text(thai_text),
            "voice": thai_voice,
            "rate": TTS_RATE,
            "volume": TTS_VOLUME,
            "pitch": TTS_PITCH,
        })

    if mode in {"english_only", "bilingual"}:
        english_text = "Attention please. This is Khlong Bang Phra Station. Thank you."
        segments.append({
            "code": "en_test",
            "label": "ภาษาอังกฤษ",
            "text": prepare_english_tts_text(english_text),
            "voice": english_voice,
            "rate": TTS_EN_RATE,
            "volume": TTS_EN_VOLUME,
            "pitch": TTS_EN_PITCH,
        })

    audio_urls = []
    audio_labels = []
    try:
        filenames = generate_audio_segments(segments)
        for segment, filename in zip(segments, filenames):
            audio_urls.append(f"/audio/{filename}")
            audio_labels.append(segment["label"])
    except FileNotFoundError:
        return jsonify({"status": "error", "message": "ยังไม่ได้ติดตั้ง edge-tts ให้รันคำสั่ง: pip install edge-tts"}), 500
    except subprocess.CalledProcessError as exc:
        detail = (exc.stderr or exc.stdout or str(exc)).strip()
        return jsonify({"status": "error", "message": f"สร้างเสียงทดสอบไม่สำเร็จ: {detail}"}), 500
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500

    if mode == "bilingual":
        preview = f"🇹🇭 {thai_text}<br><br>🇬🇧 {english_text}"
    elif mode == "english_only":
        preview = f"🇬🇧 {english_text}"
    else:
        preview = f"🇹🇭 {thai_text}"

    return jsonify({
        "status": "success",
        "audio_url": audio_urls[0] if audio_urls else "",
        "audio_urls": audio_urls,
        "audio_labels": audio_labels,
        "text_preview": preview,
        "thai_voice": thai_voice,
        "english_voice": english_voice,
    })

@app.route("/announce", methods=["POST"])
@roles_required("announcer", "admin")
def announce():
    generation_started = time.perf_counter()
    cleanup_old_audio()
    data = request.get_json(silent=True) or {}

    mode = (data.get("announce_mode") or "thai_only").strip()
    allowed_modes = {"thai_only", "english_only", "bilingual"}
    if mode not in allowed_modes:
        return jsonify({"status": "error", "message": "รูปแบบภาษาที่เลือกไม่ถูกต้อง"}), 400

    thai_text = ""
    english_text = ""
    segments = []
    thai_voice = (data.get("thai_voice") or VOICE_NAME).strip()
    if thai_voice not in THAI_VOICE_OPTIONS:
        thai_voice = VOICE_NAME

    if mode in {"thai_only", "bilingual"}:
        try:
            thai_text = build_announcement(data)
            thai_text = apply_voice_politeness(thai_text, thai_voice)
        except Exception as exc:
            return jsonify({"status": "error", "message": f"สร้างข้อความไทยไม่สำเร็จ: {exc}"}), 400

        if not thai_text:
            return jsonify({"status": "error", "message": "ไม่มีข้อความภาษาไทยสำหรับประกาศ"}), 400

        segments.append({
            "code": "th",
            "display_label": "ภาษาไทย",
            "text": thai_text,
            "voice": thai_voice,
            "rate": TTS_RATE,
            "volume": TTS_VOLUME,
            "pitch": TTS_PITCH,
            "prepare": prepare_tts_text,
        })

    if mode in {"english_only", "bilingual"}:
        try:
            english_text = build_english_announcement(data)
        except Exception as exc:
            return jsonify({"status": "error", "message": f"สร้างข้อความอังกฤษไม่สำเร็จ: {exc}"}), 400

        if not english_text:
            return jsonify({"status": "error", "message": "ไม่มีข้อความภาษาอังกฤษสำหรับประกาศ"}), 400

        segments.append({
            "code": "en",
            "display_label": "ภาษาอังกฤษ",
            "text": english_text,
            "voice": english_voice_for(thai_voice),
            "rate": TTS_EN_RATE,
            "volume": TTS_EN_VOLUME,
            "pitch": TTS_EN_PITCH,
            "prepare": prepare_english_tts_text,
        })

    audio_urls = []
    audio_labels = []

    try:
        filenames = generate_audio_segments(segments)
        for segment, filename in zip(segments, filenames):
            audio_urls.append(f"/audio/{filename}")
            audio_labels.append(segment["display_label"])

    except FileNotFoundError:
        return jsonify({
            "status": "error",
            "message": "ยังไม่ได้ติดตั้ง edge-tts ให้รันคำสั่ง: pip install edge-tts",
            "text_preview": thai_text or english_text,
        }), 500
    except subprocess.CalledProcessError as exc:
        detail = (exc.stderr or exc.stdout or str(exc)).strip()
        return jsonify({
            "status": "error",
            "message": f"สร้างเสียงไม่สำเร็จ: {detail}",
            "text_preview": thai_text or english_text,
        }), 500
    except Exception as exc:
        return jsonify({
            "status": "error",
            "message": str(exc),
            "text_preview": thai_text or english_text,
        }), 500

    if not audio_urls:
        return jsonify({
            "status": "error",
            "message": "ไม่มีไฟล์เสียงสำหรับประกาศ",
            "text_preview": thai_text or english_text,
        }), 500

    if mode == "bilingual":
        preview = f"🇹🇭 {thai_text}<br><br>🇬🇧 {english_text}"
    elif mode == "english_only":
        preview = f"🇬🇧 {english_text}"
    else:
        preview = f"🇹🇭 {thai_text}"

    return jsonify({
        "status": "success",
        "audio_url": audio_urls[0],
        "audio_urls": audio_urls,
        "audio_labels": audio_labels,
        "announce_mode": mode,
        "text_preview": preview,
        "generation_ms": int((time.perf_counter() - generation_started) * 1000),
    })

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
